"""/api/knowledge — KB management + retrieval."""
from __future__ import annotations

import io
import re

from fastapi import APIRouter, File, Request, UploadFile
from fastapi.responses import JSONResponse

from ..skills.knowledge import ingest_document, list_documents, retrieve

router = APIRouter()


def _extract_text(filename: str, buf: bytes) -> str:
    lower = filename.lower()
    if lower.endswith(".pdf"):
        try:
            from pypdf import PdfReader

            r = PdfReader(io.BytesIO(buf))
            return "\n\n".join(p.extract_text() or "" for p in r.pages)
        except Exception:
            return buf.decode("utf-8", errors="ignore")
    if re.search(r"\.xlsx?$", lower):
        try:
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(buf), data_only=True)
            chunks: list[str] = []
            for sn in wb.sheetnames:
                ws = wb[sn]
                chunks.append(f"## {sn}")
                for row in ws.iter_rows(values_only=True):
                    chunks.append(",".join("" if c is None else str(c) for c in row))
            return "\n".join(chunks)
        except Exception:
            return buf.decode("utf-8", errors="ignore")
    if lower.endswith(".docx"):
        try:
            from docx import Document

            d = Document(io.BytesIO(buf))
            return "\n".join(p.text for p in d.paragraphs if p.text)
        except Exception:
            return buf.decode("utf-8", errors="ignore")
    return buf.decode("utf-8", errors="ignore")


@router.get("/{org}")
async def list_docs(org: str) -> JSONResponse:
    return JSONResponse(content={"docs": list_documents(org)})


@router.post("/{org}")
async def upload(org: str, file: list[UploadFile] = File(...)) -> JSONResponse:
    results = []
    for f in file:
        buf = await f.read()
        text = _extract_text(f.filename, buf)
        r = await ingest_document({"org_id": org, "filename": f.filename, "text": text})
        results.append({**r, "filename": f.filename})
    return JSONResponse(content={"ok": True, "results": results})


@router.post("/{org}/retrieve")
async def retrieve_route(org: str, request: Request) -> JSONResponse:
    body = await request.json()
    return JSONResponse(
        content=await retrieve(
            {"org_id": org, "query": body["query"], "top_k": body.get("top_k") or 3}
        )
    )
