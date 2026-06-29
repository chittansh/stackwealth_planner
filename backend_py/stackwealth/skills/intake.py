"""
Universal intake — Python port of skills/intake/.

Strategy mirrors TS:
  Tier 1 deterministic — known PDF templates (AA), known XLSX templates.
  Tier 2 multimodal LLM — Claude (native PDF/image) → GPT-4o fallback.

The TS version has 8 specialized parsers; this port handles PDF, XLSX, CSV,
DOCX, MD/TXT, image, audio. The two-tier extraction strategy is preserved but
the AA-PDF deterministic regex set + xlsx-template detection are intentionally
simplified (the TS regex catalog is large; this port falls through to the LLM
on those cases — which is the same behavior the TS version uses when the
deterministic anchor doesn't match). Field paths and confidence values match
the TS contract exactly.
"""
from __future__ import annotations

import asyncio
import base64
import io
import json
import re
from datetime import datetime, timezone
from typing import Any, Optional

from .. import config
from ..logging_config import get_logger

_log = get_logger(__name__)

# Character budget for the document text shipped to the extraction LLM. The
# intake model (Haiku 4.5) has a 200K-token context, so ~200K chars (~50K
# tokens) leaves ample room and means realistic workbooks are never truncated.
# The old 80K cut silently dropped trailing sheets on large/atypical files.
_LLM_DOC_CHAR_BUDGET = 200_000


def _assemble_sheets_within_budget(sheet_blocks: list[tuple[str, list[str]]], budget: int) -> str:
    """Join per-sheet row-blocks into one document, fairly budgeted so a single
    oversized sheet can never starve (or fully drop) the others. Every sheet is
    always represented; oversized sheets are row-trimmed with an explicit marker
    rather than silently cut. Keeps extraction agnostic to how many tabs a
    workbook has and how rows are distributed across them."""
    if not sheet_blocks:
        return ""
    full = "\n".join(f"## Sheet: {sn}\n" + "\n".join(lines) for sn, lines in sheet_blocks)
    if len(full) <= budget:
        return full
    per = max(2_000, budget // len(sheet_blocks))
    parts: list[str] = []
    for sn, lines in sheet_blocks:
        kept, size = [], 0
        for ln in lines:
            if size + len(ln) + 1 > per:
                break
            kept.append(ln)
            size += len(ln) + 1
        omitted = len(lines) - len(kept)
        if omitted:
            kept.append(f"... [{omitted} more rows omitted to fit budget]")
        parts.append(f"## Sheet: {sn}\n" + "\n".join(kept))
    return "\n".join(parts)


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _empty_result(parser_used: str) -> dict[str, Any]:
    return {
        "partial_state": {},
        "evidence": [],
        "missing": [],
        "parser_used": parser_used,
    }


# ── Schema instructions for the LLM ────────────────────────────────────────


EXTRACTION_INSTRUCTIONS = """You are an Indian household financial-plan extractor. Read the document and emit a JSON object with this exact shape:

{
  "partial_state": {
    "personal_details": { "full_name"?, "date_of_birth"?, "city_of_residence"?, "city_type"? ("Metro" | "Non-metro"), "occupation"?, "retirement_age_target"?, "business_retirement_age"? (age the BUSINESS income runs to, if stated to differ from salaried retirement), "dependent_senior_parents"? (count of senior ≥60 dependent parents needing a separate health policy), "dependents"? (int OR descriptive string like "Mother (78)"), "marital_status"?, "spouse_name_and_age"?, "number_of_children"? },
    "income_details":   { "client_salary_in_hand"?, "spouse_salary_in_hand"?, "client_business_income"?, "spouse_business_income"?, "client_rental_income"?, "spouse_rental_income"?, "client_other_income"?, "spouse_other_income"?  (all monthly INR — populate every non-zero field; business owners often have 0 salary but non-zero business_income) },
    "monthly_expenses": { "household_expenses"?, "rent_or_emi"?, "groceries"?, "utilities"?, "school_fees"?, "medical"?, "insurance_premium"?, "travel_or_lifestyle"?, "other_emis"? (loan EMIs only — NOT a SIP) },
    "monthly_investments": { "mutual_fund_sip"?, "nps"?, "ppf"?, "rd"?, "direct_equity"?, "insurance_premium"?, "other"? },
    "recurring_investments": [ { "investment_type" (e.g. "Mutual Fund SIP" | "NPS" | "PPF" | "VPF" | "RD" | "Direct Equity"), "monthly_amount", "purpose" ("retirement" | "goal" | "emergency" | "general" — infer from the Remarks column), "linked_goal"? (goal name when purpose="goal", e.g. "House Purchase"), "remarks"? } ],
    "liquid_capital":   { "savings_account_balance"?, "idle_cash_for_investment"?, "fd_breakable_for_investment"? },
    "loans_liabilities": { "home_loan"? { outstanding_amount, emi, interest_rate, tenure_left }, "car_loan"?, "personal_loan"?, "credit_card_dues"? },
    "insurance_details": { "term_plan"? { "company"?, "cover_amount"?, "annual_premium"? }, "health_insurance"? { "company"?, "cover_amount"?, "annual_premium"? }, "family_floater"? { "company"?, "cover_amount"?, "annual_premium"? }, "ulip_or_endowment"? { "company"?, "cover_amount"?, "annual_premium"? } },
    "financial_goals": [ { "goal_name", "kind" ("child_education" | "child_marriage" | "retirement" | "house_purchase" | "foreign_travel" | "other"), "target_year"?, "target_amount"? (in TODAY's rupees — see Goals rules below), "is_target_in_today_money"? (bool — true if target_amount is today's cost; false ONLY if the source explicitly gives an already-inflated future-value figure with no today's cost), "inflation_assumed"? (decimal, e.g. 0.08 for 8%), "current_allocated_amount"?, "periodic_contribution"?, "contribution_frequency"? ("monthly" | "annual"), "priority"? } ],
    "mutual_funds":   [ { "fund_name", "current_value", "isin"?, "folio"?, "sip_amount"? } ],
    "equity_stocks":  [ { "stock_name", "current_value", "quantity"?, "isin"?, "long_term_or_trading"? ("long_term" | "trading") } ],
    "fixed_income":   [ { "instrument" ("FD" | "RD" | "PPF" | "EPF" | "Bonds" | "NPS" | "NSC" | "PostOffice" | "SukanyaSamriddhi" | "Other"), "invested_amount"?, "current_value"?, "maturity_date"? } ],
    "real_estate":    [ { "label" (e.g. "Self-Occupied Villa (Gurgaon)"), "kind" ("residential" | "commercial" | "land" | "other"), "current_value", "earmarked_for_sale"? (bool), "expected_appreciation_pa"? (decimal) } ],
    "gold":           [ { "label" (e.g. "Sovereign Gold Bonds"), "kind" ("physical" | "sgb" | "digital" | "jewellery"), "current_value", "held_for_investment"? (bool, default true) } ],
    "emergency_fund": { "emergency_fund_available"? (bool), "total_emergency_corpus"? (INR), "where_is_it_parked"? (string), "monthly_household_expense_for_calculation"? (INR), "months_of_cover_available"? (decimal) },
    "assumptions": {
      "persons"? [ { "name", "date_of_birth"? (DD-MM-YYYY), "life_expectancy"? (years, default 85), "retirement_age"? (AGE in years, NOT a calendar year) } ],
      "lumpsum_events"? [ { "year" (calendar year, e.g. 2031), "amount" (INR, positive = deposit/inflow, negative = withdrawal/outflow), "label" (short description like "Bonus expected", "Knee surgery", "Reverse mortgage payout", "Sale proceeds") } ]
    },
    "freedom_score_inputs": { "age"?, "monthly_income"?, "monthly_expenses"?, "monthly_emi"?, "portfolio_current_value"?, "liquid_assets_current_value"?, "equity_allocation_percent"? }
  },
  "evidence": [ { "field": "<canonical.path>", "value": <same as in partial_state>, "confidence": 0..1, "evidence_quote": "<verbatim span from source>" } ],
  "missing":  [ "<canonical.path>" ]
}

Rules:
- The document may be structured (template xlsx, bank statement, form) OR completely unstructured (a paragraph someone wrote, a screenshot, a voice transcript, a WhatsApp chat). Extract EVERY financial fact you can identify, regardless of formatting. Do NOT pattern-match on sheet names like `4A_Mutual_Funds`; rely on column-header SEMANTICS, content, and labels. The same household's data may arrive with sheets called `Mutual_Funds` or `mfs` or `Investments – MF` — analyze content, not tab names.

ANALYTICAL APPROACH — before emitting any value, REASON about it:
- **Year vs Age** (huge footgun on retirement_age):
    * "Retirement Age 60", "I'll retire at 55", "retire by age 58" → these are AGES (small int, typically 50-65). Emit as `assumptions.persons[*].retirement_age` AND `personal_details.retirement_age_target`.
    * "Retirement Year 2030", "Plan to retire in 2032", "Target year: 2035" → these are CALENDAR YEARS (typically 2025-2070). To get age: `retirement_age = retirement_year − birth_year`. Example: born 1990, "Retirement Year 2030" → retirement_age = 40. NEVER emit 2030 as retirement_age.
    * Sanity check: if you're about to emit `retirement_age > 100` you've extracted a YEAR by mistake — convert it.
    * Maturity dates, target years for goals, "born in YYYY" — these ARE legitimately calendar years, leave them as such.
- **Multi-row Personal_Details tables** (one row = one person):
    * Many templates have a TABLE like `Member | Name | Date of Birth | Marital | Retirement | Life Expectancy | City | Occupation` with rows for Client, Spouse, Child 1, Child 2, Father, Mother. Each row with a name + DOB is a person.
    * For Client + Spouse, emit BOTH into `assumptions.persons[]` with full {name, date_of_birth, retirement_age, life_expectancy}.
    * Children & dependents (Father/Mother) usually go into `personal_details.{number_of_children, dependents}` as summaries unless the household plan explicitly needs them as planning subjects — when in doubt, persons[] = primary earners only.
    * When dependents include senior parents (age ≥ 60, e.g. "Mother (78)", "Father 65 + Mother 60"), ALSO set `personal_details.dependent_senior_parents` to the count — they drive a separate health-insurance policy.
    * If the income/cash-flow source indicates business income continues to a different age than salaried retirement (e.g. business runs to 60 while employment stops at 56), set `personal_details.business_retirement_age` to that age.
- **Totals & subtotals — NEVER emit a subtotal as a field value**:
    * Subtotal / Total / Grand Total rows are AGGREGATES. They exist to verify your itemization, NOT to map into a field directly.
    * Rule: itemize EVERY component line in its proper field. Never also dump a "Subtotal" into a generic field like `household_expenses`. Doing so double-counts (subtotal + components = 2× the truth).
    * E.g. an Expenses sheet with line items {Rent 70k, Groceries 25k, Utilities 12k, School 40k, Insurance 40k, Medical 5k, Lifestyle 17k} that sum to a "Subtotal – Essential ₹2,26,000": map each line to its field (rent_or_emi=70k, groceries=25k, utilities=12k, school_fees=40k, insurance_premium=40k, medical=5k, travel_or_lifestyle=17k). The subtotal ₹2,26,000 goes NOWHERE — your itemization already sums to it.
    * Watch for sectioned tables where row #N is a section header (not a data row): "A. Essential Spends", "B. Discretionary", "I. Gross Income", "III. Net Income" are HEADERS, not values. Skip them.
    * Use the TOTAL footer as a SANITY CHECK only: sum your emitted components → if it doesn't equal the footer (within ₹2k tolerance), re-derive.
- **`household_expenses` is the RESIDUAL catchall — but NEVER for subtotals**:
    * Use `household_expenses` for line items that don't fit a more specific field. Examples: "Transport & Commute", "Other Expenses (property tax, donations, hobby)", "Maid / driver", "Pet care", "Subscriptions", "Miscellaneous", "Entertainment" if not lifestyle-y.
    * **SUM the residuals — don't pick ONE.** If the source has BOTH "Transport 10000" AND "Other Expenses 12000" AND "Maid 5000", `household_expenses = 10000+12000+5000 = 27000`. NOT 12000. NOT 10000. The single most common extraction bug here is picking the LAST residual row seen and ignoring the others — this silently drops thousands of rupees of monthly spend from the projection.
    * If a source uses "Household Expenses" as a CATEGORY HEADER with sub-line-items underneath, map the sub-line-items, not the header.
    * NEVER put a Subtotal / Total / "Essential Total" row into `household_expenses` — that doubles up with itemized components.
    * Sanity: after extraction, `sum(monthly_expenses.*)` MUST equal the TOTAL EXPENDITURE footer (within ₹2k). If it's short, you dropped some line items — add them to `household_expenses` as the residual. If it's over, you double-counted a subtotal — remove that.
- **Common line-item → field mapping cheat sheet**:
    * Rent / Maintenance / Society fees → `rent_or_emi`
    * Groceries / Living Expense / Milk / Newspaper → `groceries`
    * Utilities / Gas / Electricity / Phone / Internet / Water → `utilities`
    * Children Education / School Fees / Tuition / Uniforms → `school_fees`
    * Transport / Commute / Petrol / Car Maintenance / Cab → `household_expenses` (residual)
    * Medical / Pharmacy / Doctor visits → `medical`
    * Insurance Premium (Health, Life, Vehicle if expensed) → `insurance_premium` (sum them all)
    * Lifestyle / Restaurants / Gym / Club / Gifts / Apparel → `travel_or_lifestyle`
    * Vacation / Travel / Entertainment / Discretionary → `travel_or_lifestyle`
    * Property Tax / Donations / Sports / Hobby / Other → `household_expenses` (residual)
    * Maid / Driver / Cook / Domestic help → `household_expenses` (residual)
- **monthly_expenses is CONSUMPTION ONLY** (the single most-double-counted field):
    * IN: rent / groceries / utilities / school fees / medical / household / lifestyle / travel / insurance premium (consumption-side).
    * OUT — these are NOT monthly_expenses:
        - Salary tax deductions, TDS, professional tax → ALREADY netted out of `client_salary_in_hand` (which is post-tax in-hand). DO NOT add them again to expenses.
        - Provident Fund / EPF contributions → these are wealth-building, NOT consumption. Go in `monthly_investments` (or omit if already netted from salary).
        - SIPs / PPF / NPS contributions → `monthly_investments.*`, never expenses.
        - Loan EMIs → `monthly_expenses.other_emis` AND `loans_liabilities.*` (do not also sum into the "TOTAL expenses" figure if the source already separates loans).
    * If the source has an "Income tab Deductions" section (taxes, PF, etc.) and the salary you're reading is "Net in-hand" / "Take Home" / "Net Monthly Income", IGNORE the deductions — they're already subtracted.
    * Sanity check: `monthly_expenses` ≈ rent + groceries + utilities + lifestyle + medical + school + insurance_prem ≈ TOTAL EXPENDITURE footer (excluding loans). If your aggregate is >1.5× the footer, you've double-counted something — re-derive.
- **Semantic synonyms — map any of these to the same canonical field**:
    * "Net in-hand", "Take Home", "Net Monthly Income", "In-hand salary", "Post-tax salary" → `income_details.client_salary_in_hand` (note: this is monthly).
    * "CTC", "Gross Salary", "Annual Package", "LPA" → these are GROSS or ANNUAL. If only CTC is given, divide by 12 AND subtract ~25% tax to estimate in-hand; better: ask the user. Don't put CTC directly into client_salary_in_hand.
    * "Outstanding Principal", "Loan Outstanding", "Loan Balance" → `loans_liabilities.*.outstanding_amount`.
    * "Cover Amount", "Sum Assured", "Sum Insured" → `insurance_details.*.cover_amount`.
    * "Liquid Funds" / "Sweep FD" / "Savings A/c" → liquid_capital.* OR fixed_income depending on instrument.
- **The document may be unstructured paragraph text** — treat sentences like rows. "I earn 5 lakh a month, spend 1.5 lakh, have 30 lakh in MFs and a 1 cr flat with a 40 lakh home loan" → extract every clause: salary 500000, expenses 150000, MF total 3000000, real_estate [{label:"Flat", kind:"residential", current_value:10000000}], home_loan.outstanding_amount=4000000.


- All monetary values are monthly INR unless they're in goals / portfolios / loans (then absolute INR).
- Indian number conversion (CRITICAL — get this wrong and the entire plan is off by 10x):
    * 1 thousand / 1k = 1000
    * 1 lakh / 1 L / 1 lac = 100000 (one hundred thousand, i.e. ₹1,00,000)
    * 1 crore / 1 Cr = 10000000 (ten million, i.e. ₹1,00,00,000 — exactly 100 lakhs)
    * Worked examples — VERIFY each before emitting:
        - "2.5L" → 250000 (NOT 2500000)
        - "2.6L savings" → savings_account_balance: 260000 (NOT 2600000)
        - "1.8L in savings" → 180000 (NOT 1800000)
        - "12L" → 1200000 (NOT 12000000)
        - "28L in MFs" → 2800000
        - "50 lakh" → 5000000
        - "1.5 Cr" → 15000000 (NOT 150000000)
        - "2.5 Cr" → 25000000 (NOT 2500000)
        - "80k" → 80000
        - "12 LPA" annual → 100000 per month (divide by 12 if the schema field is monthly)
    * Sanity check before every emit: "L" / "lakh" → multiply by 100000 EXACTLY. NOT 1000000. A value like "2.6L" must produce 260000, NOT 2600000. If the original text mentions a "lakh" / "L" value and your output has more than 6 digits for amounts < 10L, you 10x'd it — re-multiply by 0.1.
    * Sanity check: 1 Cr is 100x of 1 L. If your output has 9 digits for a "crore" value, it's WRONG by 10x.
- DOB format: DD-MM-YYYY (re-format "15-Aug-1997" → "15-08-1997"). If only an age is mentioned ("im 32"), set `freedom_score_inputs.age` to that integer. Don't fabricate a DOB.
- Age: ALWAYS emit `freedom_score_inputs.age` when an age is mentioned in any form ("32 years old", "age 45", "im 28", "I'm in my 40s" → 40).
- City type: Mumbai/Delhi/Kolkata/Chennai/Bengaluru/Hyderabad/Pune/Ahmedabad = "Metro"; else "Non-metro".
- Income: write every non-zero subfield individually. Business owners often have 0 salary but non-zero business_income. For each income line decide whose it is (client vs spouse) and which kind (salary / business / rental / other). Spouse only mentioned by name with no income? Don't invent zero — omit.
- Expenses vs Investments — THE BUCKET MATTERS:
    - SIP / PPF / NPS / RD / direct-equity contributions → `monthly_investments.*` (these are wealth-building, not consumption)
    - Loan EMI of any kind → `monthly_expenses.other_emis` AND describe the loan under `loans_liabilities.*`
    - Rent, groceries, utilities, school, insurance premium, medical, travel/lifestyle → `monthly_expenses.*`
    - NEVER put a SIP under monthly_expenses — that double-counts savings as spending.
- Recurring investments — populate BOTH `monthly_investments.*` (the type-keyed aggregate) AND `recurring_investments[]` (one row per line, with PURPOSE). The recurring-investments sheet's Remarks column tells you the purpose: "For Retirement" → `purpose:"retirement"`; "For House Purchase" / "For Child Education" / any goal → `purpose:"goal"` with `linked_goal` set to that goal; emergency buffer → `purpose:"emergency"`; otherwise `purpose:"general"`. NPS / VPF / EPF are retirement vehicles → default them to `purpose:"retirement"` unless the remark says otherwise. This lets the plan net ONLY retirement-directed SIPs against the retirement corpus instead of sweeping in goal SIPs.
- Portfolio aggregation: if the user mentions a portfolio total ("my portfolio is around 12L", "I have 50L in equities"), set `freedom_score_inputs.portfolio_current_value` to that absolute INR value. If they list individual MFs/stocks, also populate the `mutual_funds[]` / `equity_stocks[]` arrays.
- Fixed income — extract EVERY non-total row in `4C_Fixed_Income[]` regardless of instrument label. Map the instrument string to one of: FD, RD, PPF, EPF, Bonds, NPS, NSC, PostOffice, SukanyaSamriddhi. Anything else (Kisan Vikas Patra, Senior Citizen Savings, debt MFs held as fixed income, Post Office Time Deposit, etc.) → "Other" with the original name preserved in optional notes. **Do NOT silently drop a row just because its instrument isn't in the canonical list — that silently strips lakhs of opening assets.** "Post Office Saving A/c" / "POSA" → "PostOffice". "National Savings Certificate" / "NSC VIII issue" → "NSC". Skip rows where current_value is blank or zero AND invested_amount is blank.
- Liquid: cash in savings / current accounts → `liquid_capital.savings_account_balance` AND `freedom_score_inputs.liquid_assets_current_value` (same total).
- Loan tenure_left: numeric YEARS only. "12 years" → 12, "3 years 6 months" → 3.5. If credit card is paid in full each month, set `tenure_left` to 0 (not a string).
- Real estate — extract every property in `real_estate[]`. The firm xlsx template has a `4D_Real_Estate` sheet with columns "Property Type | Current Market Value | Loan Outstanding | Rental Income". Map "Property Type" → `label`, "Current Market Value" → `current_value` (INR). Infer `kind`:
    * "self-occupied" / "house" / "villa" / "apartment" / "flat" → "residential"
    * "commercial" / "shop" / "office" → "commercial"
    * "plot" / "land" → "land"
    * else → "other"
  If the row also has a "Loan Outstanding" value, that's a home loan — ALSO emit a `loans_liabilities.home_loan` entry with that outstanding_amount (only ONE home_loan row aggregated across properties). If "Rental Income" is non-zero, ADD it to `income_details.client_rental_income` (or spouse if context says so). Skip "Total" / "No real estate holdings" / blank rows.
- Gold & others — extract every row in `gold[]`. The firm xlsx template has a `4E_Gold_Others` sheet with columns "Asset | Current Value". Map "Asset" → `label`. Infer `kind`:
    * "physical" / "coins" / "bars" → "physical"
    * "jewellery" → "jewellery"
    * "sovereign gold bond" / "SGB" → "sgb"
    * "digital" / "etf" → "digital"
    * else → "physical"
  Silver coins / non-gold metals also go in `gold[]` (the schema is "Gold & others" — they're treated as the same asset class for return assumptions). Skip "Total" rows.
- Emergency fund — populate `emergency_fund` (dict, NOT a list). The firm xlsx template has a `7_Emergency_Fund` sheet with rows "Emergency fund available? (Yes/No)", "Total Emergency Corpus", "Where is it parked?", "Monthly household expense (for calculation)", "Months of cover available". Map directly.
- Lumpsum events — one-off year-specific cashflows that DON'T fit goals or recurring expenses. Use `assumptions.lumpsum_events[]`. Examples:
    * "Expecting a 5 lakh bonus in 2027" → `{year: 2027, amount: 500000, label: "Bonus expected"}`
    * "Knee surgery planned in 2031 will cost about 3 lakh" → `{year: 2031, amount: -300000, label: "Knee surgery"}`
    * "Will sell apartment for 60 lakh in 2035" → `{year: 2035, amount: 6000000, label: "Apartment sale proceeds"}`
    * "Plan reverse mortgage of 20 lakh in retirement 2045" → `{year: 2045, amount: 2000000, label: "Reverse mortgage"}`
    * Sign convention: POSITIVE = cash INFLOW (deposit), NEGATIVE = cash OUTFLOW (one-off expense).
    * **Bonus expected for investment** — if `6_Liquid_Capital` (or similar) has a row labelled "Bonus Expected for Investment" with a non-empty amount, emit a lumpsum_event for the CURRENT year with that amount and label "Bonus expected for investment". If the amount is empty/blank, OMIT the event.
    * **CRITICAL — one-offs are NEVER recurring income.** `income_details.*` holds STEADY monthly streams ONLY. A one-off bonus, expected windfall, sale proceeds, maturity, or ANY amount whose label says "one time" / "one-off" / "expected" / "bonus" is a lumpsum_event — NOT income — EVEN IF it appears in (or beside) the income sheet. Concretely: an income-sheet remark like "bonus + cash : one time" or "Bonus expected - one time earlier covered in other income moved here" means that amount must go to `assumptions.lumpsum_events[]` (current year), and `client_other_income` / `spouse_other_income` must NOT include it. Do not double-count: emit it once, as a lumpsum.
    * **Do NOT** emit lumpsum_events for fixed-income instrument maturities (FD/Bond/NSC/PPF/EPF/NPS rows in `4C_Fixed_Income`). Those are already inside the FA portfolio and compound there — never re-inject them as lumpsums (it double-counts money the opening balance already holds). Do NOT set `fixed_income[].maturity_date` for the purpose of generating a lumpsum either.
    * **RM-entered manual lumpsums in a `YoY Cash Flow` tab**: if the workbook has a year-by-year cash-flow sheet with a manual "Lumpsum Further deposit / (Withdrawal)" column carrying labelled one-off entries (e.g. "Bonus Expected", "Reverse Mortgage", "Balance sale consideration", "Knee Surgery"), emit a lumpsum_event for EACH: `{year (from that row's year), amount (POSITIVE for a deposit/inflow, NEGATIVE for a one-off expense), label (the remark text)}`. **CRITICAL — do NOT capture the goal-withdrawal column** (negative values that equal a financial goal's future value, mirroring `10_Financial_Goals`): those are already modelled from the goals list and capturing them here double-counts.
  Don't confuse these with goals (those have target_year + target_amount + kind), regular monthly expenses (those go in monthly_expenses), or income (steady streams go in income_details).
- Goals: extract intent like "want to buy a 1.5cr home by 2030" → `{goal_name: "Home Purchase", kind: "house_purchase", target_year: 2030, target_amount: 15000000, is_target_in_today_money: true}`. Goal `kind` MUST be one of: child_education | child_marriage | retirement | house_purchase | foreign_travel | other.
- Goals — TODAY'S COST vs FUTURE VALUE (CRITICAL — get this wrong and every projection is inflated 2-10x):
    * `target_amount` MUST be the cost in TODAY's rupees, NOT the inflation-adjusted future value. Then set `is_target_in_today_money: true`.
    * When a structured source (e.g. an xlsx with columns) has BOTH "Today's Cost" AND "Future Value Needed" / "FV" / "Future Value", READ the today's cost — IGNORE the future value column entirely. The FV is a DERIVED computation, not an input.
    * When the source has "Inflation Assumed" or similar (e.g. "8%"), populate `inflation_assumed` as a decimal (8% → 0.08, 6% → 0.06).
    * In free-form text ("buy a 1.5cr home in 2030") the user almost always means today's money — set `is_target_in_today_money: true`.
    * The only case for `is_target_in_today_money: false` is when the source EXPLICITLY says "₹X needed in 2030 (already inflation-adjusted)" / "future value of ₹X" without giving a separate today's-cost figure. This is rare.
- Goal priority: use `essential | important | aspirational` only. Map "High" → "essential", "Medium" → "important", "Low" → "aspirational".
- Don't invent values. If a field isn't clearly present, OMIT it from partial_state and add its dotted path to "missing".
- Every field in partial_state SHOULD have a matching evidence row with a verbatim quote when possible. For derived/inferred values (e.g. FSI aggregates summed from breakdown), confidence may be lower but still emit the value.
- Output JSON only. No prose, no code fences."""


# ── Text path ──────────────────────────────────────────────────────────────


async def parse_text(args: dict[str, Any]) -> dict[str, Any]:
    text = args["text"]
    source_type = args.get("source_type", "user")
    filename = args.get("filename")
    return await _llm_extract(text=text, source_type=source_type, filename=filename, parser_label=f"text:{source_type}")


# ── LLM extraction ─────────────────────────────────────────────────────────


def _stamp_evidence(rows: list[dict], source_type: str, source_file: Optional[str], parser_tier: str) -> list[dict]:
    out = []
    for r in rows:
        if not isinstance(r, dict):
            continue
        out.append(
            {
                "field": r.get("field"),
                "value": r.get("value"),
                "source_file": source_file,
                "source_type": source_type,
                "parser_tier": parser_tier,
                "confidence": float(r.get("confidence") or 0.6),
                "evidence_quote": r.get("evidence_quote"),
                "page_or_sheet": r.get("page_or_sheet"),
                "timestamp": _now(),
            }
        )
    return out


def _claude_client() -> Optional[Any]:
    if not config.ANTHROPIC_API_KEY:
        return None
    try:
        from anthropic import Anthropic

        return Anthropic(api_key=config.ANTHROPIC_API_KEY)
    except Exception:
        return None


def _openai_client() -> Optional[Any]:
    if not config.OPENAI_API_KEY:
        return None
    try:
        from openai import OpenAI

        return OpenAI(api_key=config.OPENAI_API_KEY)
    except Exception:
        return None


def _try_parse_json(s: str) -> Optional[dict]:
    s = s.strip()
    # Strip code fences if model added them.
    s = re.sub(r"^```(?:json)?\s*|\s*```$", "", s, flags=re.IGNORECASE | re.MULTILINE).strip()
    try:
        return json.loads(s)
    except Exception:
        m = re.search(r"\{.*\}", s, flags=re.DOTALL)
        if m:
            try:
                return json.loads(m.group(0))
            except Exception:
                return None
    return None


async def _llm_extract(
    *,
    text: Optional[str] = None,
    pdf_bytes: Optional[bytes] = None,
    image_bytes: Optional[bytes] = None,
    image_mime: Optional[str] = None,
    source_type: str,
    filename: Optional[str],
    parser_label: str,
) -> dict[str, Any]:
    """Try Claude first, GPT-4o as fallback, both with JSON mode."""
    out: Optional[dict] = None

    claude = _claude_client()
    if claude is not None:
        try:
            # User content holds only the document being parsed. The big
            # EXTRACTION_INSTRUCTIONS prompt (~6KB) is in the system
            # parameter with cache_control so Anthropic skips re-processing
            # it on subsequent uploads within the cache TTL (~5 min).
            # Zero impact on accuracy — same prompt, same model, same
            # inputs; the API just doesn't re-prefill the schema portion.
            # First call: full latency. Every subsequent upload during the
            # RM's session: 50–90% faster prefill.
            user_blocks: list[dict] = []
            if text:
                user_blocks.append({"type": "text", "text": f"# Document\n\n{text[:_LLM_DOC_CHAR_BUDGET]}"})
            if pdf_bytes:
                user_blocks.append(
                    {
                        "type": "document",
                        "source": {
                            "type": "base64",
                            "media_type": "application/pdf",
                            "data": base64.b64encode(pdf_bytes).decode(),
                        },
                    }
                )
            if image_bytes:
                user_blocks.append(
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": image_mime or "image/jpeg",
                            "data": base64.b64encode(image_bytes).decode(),
                        },
                    }
                )
            # Some payloads (audio transcripts that produced empty text)
            # might have no user blocks; fall through with a stub so the
            # API doesn't 400.
            if not user_blocks:
                user_blocks.append({"type": "text", "text": "(no document content)"})
            # The Anthropic SDK's `messages.create` is BLOCKING. Calling
            # it directly from an async handler pins the whole asyncio
            # event loop for the duration of the LLM call (~90s for a
            # full xlsx). During that window even /health can't respond,
            # which makes Fly's healthcheck time out, drop the machine
            # from the load balancer, and surface "not listening on
            # 0.0.0.0:4000" to fly-doctor. asyncio.to_thread offloads
            # the blocking call to a worker thread so the event loop
            # keeps serving other requests.
            resp = await asyncio.to_thread(
                claude.messages.create,
                model=config.INTAKE_MODEL or "claude-haiku-4-5-20251001",
                # Firm xlsx templates produce ~8–14k tokens of JSON; 4096
                # truncated mid-document and the parser silently fell through
                # to the no-llm fallback. 16k leaves headroom for the
                # densest workbooks.
                max_tokens=16384,
                temperature=0,
                system=[
                    {
                        "type": "text",
                        "text": EXTRACTION_INSTRUCTIONS,
                        # Cache the schema for repeat uploads. Anthropic
                        # requires a minimum prefix size for caching
                        # (~1024 tokens for Haiku); EXTRACTION_INSTRUCTIONS
                        # is well above that.
                        "cache_control": {"type": "ephemeral"},
                    }
                ],
                messages=[{"role": "user", "content": user_blocks}],
            )
            raw = "".join(b.text for b in resp.content if hasattr(b, "text"))
            if getattr(resp, "stop_reason", None) == "max_tokens":
                _log.warning(
                    "intake.claude.truncated",
                    extra={"output_tokens": getattr(resp.usage, "output_tokens", None),
                           "category": "llm"},
                )
            out = _try_parse_json(raw)
        except Exception:
            _log.error("intake.claude.failed", extra={"category": "llm"}, exc_info=True)

    if out is None:
        oa = _openai_client()
        if oa is not None:
            try:
                content: list[dict] = [{"type": "text", "text": EXTRACTION_INSTRUCTIONS}]
                if text:
                    content.append({"type": "text", "text": f"\n\n# Document\n\n{text[:_LLM_DOC_CHAR_BUDGET]}"})
                if image_bytes:
                    b64 = base64.b64encode(image_bytes).decode()
                    content.append(
                        {
                            "type": "image_url",
                            "image_url": {"url": f"data:{image_mime or 'image/jpeg'};base64,{b64}"},
                        }
                    )
                # Same event-loop concern as the Claude path above.
                resp = await asyncio.to_thread(
                    oa.chat.completions.create,
                    model="gpt-4o",
                    response_format={"type": "json_object"},
                    temperature=0,
                    messages=[{"role": "user", "content": content}],
                )
                raw = resp.choices[0].message.content or ""
                out = _try_parse_json(raw)
            except Exception:
                _log.error("intake.openai.failed", extra={"category": "llm"}, exc_info=True)

    if out is None:
        return _empty_result(f"{parser_label}:no-llm")

    partial = out.get("partial_state") or {}
    evidence = _stamp_evidence(out.get("evidence") or [], source_type, filename, "llm")
    missing = out.get("missing") or []
    return {
        "partial_state": partial,
        "evidence": evidence,
        "missing": missing,
        "parser_used": parser_label,
    }


# ── PDF / XLSX / CSV / DOCX / image / audio dispatch ───────────────────────


async def _parse_pdf(buf: bytes, filename: str) -> dict[str, Any]:
    """Try Claude with native PDF doc-block; if no LLM available, fall back to
    extracted text."""
    out = await _llm_extract(
        pdf_bytes=buf, source_type="pdf_generic", filename=filename, parser_label="pdfGeneric:claude"
    )
    if out["evidence"]:
        return out
    # Fallback: extract text and try again.
    try:
        from pypdf import PdfReader

        reader = PdfReader(io.BytesIO(buf))
        text = "\n\n".join(page.extract_text() or "" for page in reader.pages)
    except Exception as e:
        return {
            **_empty_result("pdfGeneric:failed"),
            "missing": [str(e)],
        }
    return await _llm_extract(
        text=text, source_type="pdf_generic", filename=filename, parser_label="pdfGeneric:textFallback"
    )


def _goal_kind_from_name(name: str) -> str:
    """Map a free-text goal label to a canonical GoalKind."""
    n = (name or "").lower()
    if any(k in n for k in ("education", "college", "school", "tuition", "study")):
        return "child_education"
    if any(k in n for k in ("marriage", "wedding")):
        return "child_marriage"
    if any(k in n for k in ("retire", "fire", "pension")):
        return "retirement"
    if any(k in n for k in ("house", "home", "property", "flat", "apartment", "villa", "plot", "real estate")):
        return "house_purchase"
    if any(k in n for k in ("travel", "foreign", "vacation", "trip", "holiday", "tour")):
        return "foreign_travel"
    return "other"


_INR_AMT_RE = re.compile(r"([0-9][0-9,]*\.?[0-9]*)\s*(cr|crore|crores|l|lac|lakh|lakhs|k|thousand)?", re.I)


def _parse_inr_amount(v: Any) -> Optional[float]:
    """Parse an INR amount cell — number, or a string like '2 Crore', '50L',
    '₹6 Crore', '1,50,000'. Returns None for blank/zero/unparseable."""
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v) if v else None
    if not isinstance(v, str):
        return None
    s = v.strip().lower().replace("₹", "").replace("rs.", "").replace("rs", "").strip()
    if not s:
        return None
    m = _INR_AMT_RE.search(s)
    if not m:
        return None
    try:
        num = float(m.group(1).replace(",", ""))
    except ValueError:
        return None
    unit = (m.group(2) or "").lower()
    if unit in ("cr", "crore", "crores"):
        num *= 1e7
    elif unit in ("l", "lac", "lakh", "lakhs"):
        num *= 1e5
    elif unit in ("k", "thousand"):
        num *= 1e3
    return num if num > 0 else None


def _parse_goals_sheet(wb) -> list[dict]:
    """Deterministically capture EVERY named goal row from the financial-goals
    sheet — INCLUDING name-only rows the LLM drops because they carry no amount
    or year. Each row becomes a goal with `kind` inferred from the name;
    target_year and target_amount are filled when present and left null
    otherwise, so a goal the RM has only named (e.g. 'House Purchase') still
    appears on the canvas for them to quantify, instead of silently vanishing."""
    sheet = None
    for sn in wb.sheetnames:
        if "goal" in sn.lower():
            sheet = wb[sn]
            break
    if sheet is None:
        return []
    rows = list(sheet.iter_rows(values_only=True))
    hdr_idx = None
    col: dict[str, int] = {}
    for i, row in enumerate(rows):
        low = [str(c).lower().strip() if c is not None else "" for c in row]
        has_goal = any(c == "goal" or c.startswith("goal") for c in low)
        has_field = any(("year" in c) or ("future value" in c) or ("cost" in c) for c in low)
        if has_goal and has_field:
            hdr_idx = i
            for j, c in enumerate(low):
                if c.startswith("goal") and "name" not in col:
                    col["name"] = j
                elif ("target year" in c) or c == "year":
                    col["year"] = j
                elif ("today" in c and "cost" in c) or c == "cost":
                    col["today"] = j
                elif "future value" in c:
                    col["fv"] = j
            break
    if hdr_idx is None or "name" not in col:
        return []

    out: list[dict] = []
    for row in rows[hdr_idx + 1:]:
        ni = col["name"]
        name = row[ni] if ni < len(row) else None
        if not isinstance(name, str):
            continue
        name = name.strip()
        nl = name.lower()
        # Skip header/footer/note rows that aren't real goals (e.g. a
        # "Note - Total cost in Per Annum" trailer under the goal list).
        if (not name or nl in ("total", "goal", "goals")
                or nl.startswith("note") or "total cost" in nl or "per annum" in nl):
            continue
        goal: dict[str, Any] = {"goal_name": name, "kind": _goal_kind_from_name(name)}
        yr = row[col["year"]] if col.get("year") is not None and col["year"] < len(row) else None
        if isinstance(yr, (int, float)) and not isinstance(yr, bool) and 2000 <= yr <= 2100:
            goal["target_year"] = int(yr)
        today = row[col["today"]] if col.get("today") is not None and col["today"] < len(row) else None
        fv = row[col["fv"]] if col.get("fv") is not None and col["fv"] < len(row) else None
        amt_today = _parse_inr_amount(today)
        amt_fv = _parse_inr_amount(fv)
        if amt_today:
            goal["target_amount"] = amt_today
            goal["is_target_in_today_money"] = True
        elif amt_fv:
            goal["target_amount"] = amt_fv
            goal["is_target_in_today_money"] = False
        out.append(goal)
    return out


def _merge_goal_rows(partial_state: dict, goal_rows: list[dict]) -> None:
    """Append any named goal the deterministic sheet parse found that the LLM
    didn't already extract (matched case-insensitively by name). Never drops or
    overwrites an LLM-parsed goal — only fills in the ones it missed."""
    if not goal_rows:
        return
    existing = partial_state.get("financial_goals") or []
    have = {
        str(g.get("goal_name", "")).strip().lower()
        for g in existing if isinstance(g, dict)
    }
    merged = list(existing)
    for g in goal_rows:
        if str(g.get("goal_name", "")).strip().lower() not in have:
            merged.append(g)
    partial_state["financial_goals"] = merged


def _row_is_blank(row: tuple) -> bool:
    """A row is blank if every cell is None or an empty-string after
    stripping. Excel templates have hundreds of trailing empty rows
    that we don't want to ship to the LLM — they eat the 80K-char
    truncation budget that should go to actual data."""
    for c in row:
        if c is None:
            continue
        if isinstance(c, str) and not c.strip():
            continue
        return False
    return True


def _parse_yoy_manual_inputs(wb) -> dict[str, Any]:
    """Deterministically read the firm 'YoY Cash Flow' tab's RM-MANUAL columns —
    the judgement entries that aren't derivable from the input tabs and that the
    LLM extracts unreliably:
      • 'Lumpsum Further deposit / (Withdrawal)' column + 'Remarks' → lumpsum_events
        (Bonus, Reverse Mortgage, Balance sale, Knee Surgery, ...). The adjacent
        'Major Withdrawals' column (=-SUMIF on the goals list) is SKIPPED — those
        are already modelled from 10_Financial_Goals.
      • last year the 'Income from business' column is non-zero → business income
        runs to that age (business_retirement_age), which can exceed the salaried
        retirement age.
    """
    sheet = None
    for sn in wb.sheetnames:
        s = sn.lower()
        if "yoy" in s or "cash flow" in s or "cashflow" in s:
            sheet = wb[sn]
            break
    if sheet is None:
        return {}
    max_c = sheet.max_column
    # Locate the header row + columns by header text.
    hdr_row = None
    col = {}
    for r in range(1, min(12, sheet.max_row) + 1):
        vals = [sheet.cell(row=r, column=c).value for c in range(1, max_c + 1)]
        joined = " ".join(str(v).lower() for v in vals if v)
        if "lumpsum" in joined and "business" in joined:
            hdr_row = r
            for c, v in enumerate(vals, start=1):
                t = str(v or "").lower().strip()
                if "lumpsum" in t and "further" in t:
                    col["lumpsum"] = c
                elif t == "remarks":
                    col["remarks"] = c
                elif "business" in t:
                    col["business"] = c
                elif "age" in t and "primary" in t:
                    col["age"] = c
                elif "loan" in t and "repay" in t:
                    col["loan"] = c
            break
    if not hdr_row or "lumpsum" not in col:
        return {}
    # Year column = first column carrying 4-digit years in the data rows. The
    # row directly under the header is often a sub-header (growth rates), so scan
    # the first several rows.
    year_col = None
    for c in range(1, max_c + 1):
        for rr in range(hdr_row + 1, min(hdr_row + 9, sheet.max_row) + 1):
            v = sheet.cell(row=rr, column=c).value
            if isinstance(v, (int, float)) and 2000 <= v <= 2100:
                year_col = c
                break
        if year_col:
            break
    if not year_col:
        return {}

    events: list[dict] = []
    last_biz_age = None
    loan_years = 0
    for r in range(hdr_row + 1, sheet.max_row + 1):
        yr = sheet.cell(row=r, column=year_col).value
        if not isinstance(yr, (int, float)) or not (2000 <= yr <= 2100):
            continue
        yr = int(yr)
        lump = sheet.cell(row=r, column=col["lumpsum"]).value if col.get("lumpsum") else None
        label = sheet.cell(row=r, column=col["remarks"]).value if col.get("remarks") else None
        if isinstance(lump, (int, float)) and lump != 0:
            events.append({"year": yr, "amount": float(lump), "label": str(label or "One-time event")})
        biz = sheet.cell(row=r, column=col["business"]).value if col.get("business") else None
        if isinstance(biz, (int, float)) and biz > 0:
            a = sheet.cell(row=r, column=col["age"]).value if col.get("age") else None
            if isinstance(a, (int, float)):
                last_biz_age = int(round(a))
        loan = sheet.cell(row=r, column=col["loan"]).value if col.get("loan") else None
        if isinstance(loan, (int, float)) and loan > 0:
            loan_years += 1

    out: dict[str, Any] = {}
    if events:
        out["lumpsum_events"] = events
    if last_biz_age is not None:
        out["business_retirement_age"] = last_biz_age
    if loan_years > 0:
        out["loan_years"] = loan_years
    return out


def _parse_retirement_inputs(wb) -> dict[str, Any]:
    """Deterministically read the firm 'Retirement Plan' tab's RM-manual inputs,
    which OVERRIDE the personal-details tab for the retirement corpus and step-up:
      • self / spouse life expectancy (E12 / E13 — often more conservative than
        the personal-details LE),
      • one-time post-retirement spend + its horizon (E26 / E27),
      • step-up starting annual contribution (E54) and step-up rate (F51),
      • current corpus already allocated to retirement (H53).
    """
    ws = None
    for sn in wb.sheetnames:
        s = sn.lower()
        if "retirement plan" in s and "case study" not in s:
            ws = wb[sn]
            break
    if ws is None:
        return {}

    def col_val(r, c):
        return ws.cell(row=r, column=c).value

    out: dict[str, Any] = {}
    les: list[float] = []
    for r in range(1, min(ws.max_row, 80) + 1):
        b = str(col_val(r, 2) or "").lower()
        e = col_val(r, 5)  # column E (the value column)
        if "life expectancy" in b and isinstance(e, (int, float)):
            les.append(float(e))
        elif "one time spend post" in b and isinstance(e, (int, float)):
            out["one_time_spend"] = float(e)
        elif "time for above spend" in b and isinstance(e, (int, float)):
            out["one_time_years"] = int(round(float(e)))
        elif "annual contribution to fv" in b:
            if isinstance(e, (int, float)):
                out["stepup_start_annual"] = float(e)
            h = col_val(r - 1, 8)  # H of the "One Time corpus to FV" row above
            if isinstance(h, (int, float)):
                out["corpus_allocated"] = float(h)
    # First LE = self, second = spouse (sheet lists Mr then Mrs).
    if len(les) >= 1:
        out["self_life_expectancy"] = int(round(les[0]))
    if len(les) >= 2:
        out["spouse_life_expectancy"] = int(round(les[1]))
    # Step-up rate sits in F near the Section-3 header ("Increase p.a").
    for r in range(45, min(ws.max_row, 55) + 1):
        f = col_val(r, 6)
        if isinstance(f, (int, float)) and 0 < f < 1:
            out["stepup_rate"] = float(f)
            break
    return out


# Labels that mark the end of the holdings list, not a holding. The firm's
# templates append a grand-total row plus a "Summary by Tag" block, and some
# RM-prepared sheets (e.g. Dhruv's) restate the same total as a
# re-classification ("Considered for playing in Equity" + "Available for sale")
# — counting those rows double-counts the portfolio. Once a row's name matches,
# everything below it is ignored.
_EQUITY_STOP_LABELS = (
    "total", "subtotal", "grand total", "summary", "by tag",
    "considered for playing", "available for sale",
)


def _parse_equity_stocks(wb) -> list[dict] | None:
    """Deterministically read the firm '4B_Equity_Stocks' tab. Column layouts
    vary across the firm's templates, so the value column is mislabelled in two
    different ways:
      • some templates shift the headers, so the 'Current Price' column actually
        holds each holding's TOTAL current value while 'Current Value' holds the
        strength tag;
      • others label the per-share price 'Current Price' and the total
        'Current Value'.
    The LLM intermittently computes Quantity × per-share price and inflates
    equity 100×. We read the value column directly — preferring the
    'Current Value' header when its data is numeric, else 'Current Price' — and
    never multiply by quantity. We also stop at the grand-total / summary /
    re-classification rows below the holdings (see _EQUITY_STOP_LABELS) so a
    restated total isn't summed on top of the holdings."""
    ws = None
    for sn in wb.sheetnames:
        s = sn.lower()
        if "equity" in s or s.strip().startswith("4b"):
            ws = wb[sn]
            break
    if ws is None:
        return None
    max_c = ws.max_column
    hdr = None
    name_c = value_c = price_c = None
    cand: list[int] = []
    for r in range(1, min(8, ws.max_row) + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, max_c + 1)]
        joined = " ".join(str(v).lower() for v in vals if v)
        if "stock name" in joined or ("stock" in joined and ("current value" in joined or "current price" in joined)):
            hdr = r
            for c, v in enumerate(vals, start=1):
                t = str(v or "").lower().strip()
                if "stock name" in t or t == "stock name":
                    name_c = c
                if "current value" in t:
                    value_c = value_c or c
                    cand.append(c)
                elif "current price" in t:
                    price_c = price_c or c
                    cand.append(c)
            break
    if not hdr or not name_c or not cand:
        return None

    def _col_is_numeric(c: int) -> bool:
        for rr in range(hdr + 1, min(hdr + 6, ws.max_row) + 1):
            v = ws.cell(row=rr, column=c).value
            if isinstance(v, (int, float)) and v > 0:
                return True
        return False

    # Prefer the 'Current Value' column; fall back to 'Current Price' (which on
    # shifted-header templates is the real total). Last resort: any numeric
    # candidate, left to right.
    val_c = None
    for c in (value_c, price_c, *cand):
        if c and _col_is_numeric(c):
            val_c = c
            break
    if not val_c:
        return None
    out: list[dict] = []
    for r in range(hdr + 1, ws.max_row + 1):
        name = ws.cell(row=r, column=name_c).value
        val = ws.cell(row=r, column=val_c).value
        # An unlabelled numeric row is the column's grand total — holdings end
        # here (everything below is a total / summary / re-classification block).
        if not (isinstance(name, str) and name.strip()):
            if isinstance(val, (int, float)) and val > 0:
                break
            continue
        nl = name.strip().lower()
        if any(lbl in nl for lbl in _EQUITY_STOP_LABELS):
            break
        if isinstance(val, (int, float)) and val > 0:
            out.append({"stock_name": name.strip(), "current_value": float(val)})
    if out:
        _log.info(
            "intake.equity.parsed",
            extra={"holdings": len(out),
                   "total_value": round(sum(h["current_value"] for h in out)),
                   "value_col": val_c, "category": "intake"},
        )
    return out or None


def _parse_recurring_investments(wb) -> list[dict] | None:
    """Deterministically read the firm '5_Recurring_Investments' tab so no
    monthly contribution is dropped (the LLM intermittently misses rows like
    EPF/VPF). Tags purpose so retirement instruments (NPS, VPF, EPF, PPF,
    provident/pension) feed the retirement SIP; MF/RD/direct-equity feed goals.
    The 'For Retirement' / 'For <goal>' remark wins when present."""
    ws = None
    for sn in wb.sheetnames:
        if "recurring" in sn.lower() or sn.strip().startswith("5_"):
            ws = wb[sn]
            break
    if ws is None:
        return None
    # Header row: find columns for type / amount / remarks.
    hdr = type_c = amt_c = rem_c = None
    for r in range(1, min(8, ws.max_row) + 1):
        vals = [str(ws.cell(row=r, column=c).value or "").lower() for c in range(1, ws.max_column + 1)]
        joined = " ".join(vals)
        if ("investment" in joined or "type" in joined) and ("amount" in joined or "monthly" in joined):
            hdr = r
            for c, v in enumerate(vals, start=1):
                if "type" in v or "investment" in v:
                    type_c = type_c or c
                elif "amount" in v or "monthly" in v:
                    amt_c = c
                elif "remark" in v or "comment" in v:
                    rem_c = c
            break
    if not hdr or not type_c or not amt_c:
        return None

    retire_kw = ("nps", "vpf", "epf", "ppf", "provident", "pension", "sukanya", "superannuation")
    goal_kw = ("house", "education", "college", "car", "travel", "vacation", "marriage", "wedding", "child", "goal")
    out: list[dict] = []
    for r in range(hdr + 1, ws.max_row + 1):
        name = ws.cell(row=r, column=type_c).value
        amt = ws.cell(row=r, column=amt_c).value
        if not isinstance(name, str) or not name.strip():
            continue
        nlow = name.strip().lower()
        if "total" in nlow or "insurance" in nlow:
            continue
        if not isinstance(amt, (int, float)) or amt <= 0:
            continue
        rem = str((ws.cell(row=r, column=rem_c).value if rem_c else "") or "")
        rl = rem.lower()
        if "retire" in rl:
            purpose = "retirement"
        elif any(k in rl for k in goal_kw):
            purpose = "goal"
        elif any(k in nlow for k in retire_kw):
            purpose = "retirement"
        elif any(k in nlow for k in ("mutual", "mf", "sip", "equity", "rd", "stock")):
            purpose = "goal"
        else:
            purpose = "general"
        out.append({"investment_type": name.strip(), "monthly_amount": float(amt),
                    "purpose": purpose, "remarks": rem or None})

    # Provident Fund / EPF contribution often sits in the INCOME tab as a
    # mandatory salary deduction (it reduces net pay) — but it's a monthly
    # contribution to a retirement instrument, so it must count toward the
    # retirement SIP. Capture it from the income sheet's deductions section.
    # Only skip if the MANDATORY EPF / Provident Fund is already a row — VPF
    # (voluntary PF) is a separate contribution and must not suppress it.
    have_pf = any("provident" in (r["investment_type"] or "").lower()
                  or (r["investment_type"] or "").lower().strip() == "epf"
                  for r in out)
    if not have_pf:
        for sn in wb.sheetnames:
            s = sn.lower()
            if not ("income" in s or s.strip().startswith("2_")):
                continue
            iws = wb[sn]
            for r in range(1, iws.max_row + 1):
                for c in range(1, iws.max_column + 1):
                    cell = iws.cell(row=r, column=c).value
                    if isinstance(cell, str) and ("provident fund" in cell.lower()
                                                  or cell.strip().lower() in ("epf", "vpf")):
                        nums = [iws.cell(row=r, column=cc).value for cc in range(c + 1, iws.max_column + 1)]
                        amt = max((n for n in nums if isinstance(n, (int, float)) and n > 0), default=0)
                        if amt > 0:
                            out.append({"investment_type": "EPF / Provident Fund", "monthly_amount": float(amt),
                                        "purpose": "retirement", "remarks": "Provident Fund contribution (salary deduction)"})
                        break
                else:
                    continue
                break
            break
    return out or None


async def _parse_xlsx(buf: bytes, filename: str) -> dict[str, Any]:
    expense_footer_total: float | None = None
    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(buf), data_only=True)
        # Deterministically capture the YoY tab's RM-manual columns (lumpsums +
        # business-income cut-off) before the text chunking, so they survive
        # regardless of what the LLM extracts.
        try:
            yoy_manual = _parse_yoy_manual_inputs(wb)
        except Exception:
            yoy_manual = {}
        try:
            equity_rows = _parse_equity_stocks(wb)
        except Exception:
            equity_rows = None
        try:
            retirement_inputs = _parse_retirement_inputs(wb)
        except Exception:
            retirement_inputs = {}
        try:
            recurring_rows = _parse_recurring_investments(wb)
        except Exception:
            recurring_rows = None
        try:
            goal_rows = _parse_goals_sheet(wb)
        except Exception:
            goal_rows = []
        sheet_blocks: list[tuple[str, list[str]]] = []
        for sn in wb.sheetnames:
            # Skip COMPUTED / projection tabs — they hold derived figures
            # (future balances, corpus, FV, networth roll-forwards) that the LLM
            # otherwise mistakes for input asset values, inflating the opening
            # pool 10-100×. The deterministic parsers (_parse_yoy_manual_inputs,
            # the expense-footer reconcile) read these tabs directly when needed;
            # the LLM only ever sees the raw INPUT sheets.
            _s = sn.lower()
            if any(p in _s for p in (
                "yoy", "cash flow", "cashflow", "retirement plan", "assumption",
                "insurance computation", "debt mgt", "tax planning", "checks",
                "list of tab", "networth", "inc exp", "asset returns", "case study",
            )):
                continue
            ws = wb[sn]
            # Only include sheets with at least one non-blank row, and
            # skip blank rows within a sheet. The firm's templates pad
            # every tab to 1000+ rows; without filtering, the goals
            # sheet alone consumed ~44KB of the 80KB LLM input budget
            # for empty cells, starving the other tabs.
            sheet_lines: list[str] = []
            for row in ws.iter_rows(values_only=True):
                if _row_is_blank(row):
                    continue
                sheet_lines.append(",".join("" if c is None else str(c) for c in row))
                # Capture the "Total Expenditure" footer on the expenses
                # sheet so we can reconcile against the LLM's extraction
                # downstream. The firm template's expense sheet name is
                # exactly "3_Expenses " (trailing space) but we match any
                # sheet whose name starts with "3_" or contains "expense".
                if expense_footer_total is None and ("expense" in sn.lower() or sn.lower().startswith("3_")):
                    label = next((str(c) for c in row if isinstance(c, str)), "").lower()
                    if "total expenditure" in label or "total expenses" in label:
                        nums = [c for c in row if isinstance(c, (int, float)) and c > 1000]
                        if nums:
                            expense_footer_total = float(max(nums))
            if not sheet_lines:
                continue
            sheet_blocks.append((sn, sheet_lines))
        text = _assemble_sheets_within_budget(sheet_blocks, _LLM_DOC_CHAR_BUDGET)
        _log.info(
            "intake.xlsx.dump",
            extra={"sheets": len(sheet_blocks), "chars": len(text),
                   "budget": _LLM_DOC_CHAR_BUDGET,
                   "trimmed": len(text) >= _LLM_DOC_CHAR_BUDGET, "category": "intake"},
        )
    except Exception as e:
        return {**_empty_result("xlsx:failed"), "missing": [str(e)]}
    result = await _llm_extract(
        text=text, source_type="xlsx", filename=filename, parser_label="xlsx:llm"
    )

    # Deterministic capture of the YoY tab's RM-manual columns — authoritative
    # over the LLM (which extracts these unreliably). Reverse mortgages / asset
    # sales / bonuses go to assumptions.lumpsum_events; the business-income
    # cut-off age to personal_details.business_retirement_age.
    # Deterministic equity holdings override the LLM (mislabelled-column safety).
    if equity_rows and result.get("partial_state") is not None:
        result["partial_state"]["equity_stocks"] = equity_rows

    # Deterministic recurring investments — no monthly contribution dropped, and
    # NPS/VPF/EPF/PPF correctly tagged retirement so they feed the retirement SIP.
    if recurring_rows and result.get("partial_state") is not None:
        result["partial_state"]["recurring_investments"] = recurring_rows

    # Retirement Plan tab's RM-manual inputs (life expectancies, one-time spend,
    # step-up start/rate, allocated corpus) → authoritative for the corpus + table.
    if retirement_inputs and result.get("partial_state") is not None:
        result["partial_state"]["retirement_plan_inputs"] = retirement_inputs

    # Deterministic goals capture — ensure EVERY named goal row reaches the plan,
    # even the name-only ones the LLM omits for having no amount/year. Only adds
    # goals the LLM missed; never overwrites an LLM-parsed goal.
    if goal_rows and result.get("partial_state") is not None:
        _merge_goal_rows(result["partial_state"], goal_rows)

    if yoy_manual and result.get("partial_state") is not None:
        ps = result["partial_state"]
        if yoy_manual.get("lumpsum_events"):
            asn = dict(ps.get("assumptions") or {})
            asn["lumpsum_events"] = yoy_manual["lumpsum_events"]
            ps["assumptions"] = asn
        if yoy_manual.get("business_retirement_age"):
            pdt = dict(ps.get("personal_details") or {})
            pdt["business_retirement_age"] = yoy_manual["business_retirement_age"]
            ps["personal_details"] = pdt
        if yoy_manual.get("loan_years"):
            # The YoY 'Loan Repayments' column is the firm's "Actual" EMI schedule
            # — authoritative over the Loans tab's stated tenure. Set the home
            # loan's remaining tenure to the number of years EMIs actually run.
            ll = dict(ps.get("loans_liabilities") or {})
            hl = dict(ll.get("home_loan") or {})
            hl["tenure_left"] = yoy_manual["loan_years"]
            ll["home_loan"] = hl
            ps["loans_liabilities"] = ll

    # Deterministic safety net: if the workbook's "Total Expenditure" footer
    # is materially higher than the sum of LLM-extracted monthly_expenses,
    # the LLM dropped a line item (most commonly "Transport & Commute" or
    # a second "Other Expenses" row). Auto-balance by adding the missing
    # delta to household_expenses — the firm's residual catch-all. This
    # prevents silent under-projection of lifetime expenses by 10-15%.
    if expense_footer_total and result.get("partial_state"):
        me = result["partial_state"].get("monthly_expenses") or {}
        if isinstance(me, dict):
            extracted = sum(v for v in me.values() if isinstance(v, (int, float)))
            shortfall = expense_footer_total - extracted
            # Only patch if the shortfall is real (>₹2k) and plausible
            # (<half of footer — beyond that, the LLM has bigger problems).
            if 2000 < shortfall < expense_footer_total / 2:
                me["household_expenses"] = float(me.get("household_expenses") or 0) + shortfall
                result["partial_state"]["monthly_expenses"] = me
                ev = result.get("evidence") or []
                ev.append({
                    "field_path": "monthly_expenses.household_expenses",
                    "value": me["household_expenses"],
                    "source_text": f"Auto-reconciled to match Total Expenditure footer ₹{expense_footer_total:,.0f}",
                    "confidence": 0.6,
                    "source_type": "xlsx",
                    "source_file": filename or "",
                    "extractor": "xlsx_footer_reconciler",
                })
                result["evidence"] = ev
                _log.info(
                    "intake.expense.reconciled",
                    extra={"llm_sum": round(extracted), "footer_total": round(expense_footer_total),
                           "added_to_household": round(shortfall), "category": "intake"},
                )
    return result


async def _parse_csv(buf: bytes, filename: str) -> dict[str, Any]:
    text = buf.decode("utf-8", errors="ignore")
    return await _llm_extract(
        text=text, source_type="csv", filename=filename, parser_label="csv:llm"
    )


async def _parse_docx(buf: bytes, filename: str) -> dict[str, Any]:
    try:
        from docx import Document

        doc = Document(io.BytesIO(buf))
        text = "\n".join(p.text for p in doc.paragraphs if p.text)
    except Exception as e:
        return {**_empty_result("docx:failed"), "missing": [str(e)]}
    return await _llm_extract(
        text=text, source_type="docx", filename=filename, parser_label="docx:llm"
    )


async def _parse_image(buf: bytes, filename: str, mime: str) -> dict[str, Any]:
    return await _llm_extract(
        image_bytes=buf,
        image_mime=mime,
        source_type="image",
        filename=filename,
        parser_label="image:claude",
    )


async def _parse_audio(buf: bytes, filename: str, mime: str) -> dict[str, Any]:
    """Whisper transcribe → parse_text."""
    oa = _openai_client()
    if oa is None:
        return {**_empty_result("audio:no-openai"), "missing": ["OPENAI_API_KEY required for audio"]}
    try:
        f = io.BytesIO(buf)
        f.name = filename
        r = oa.audio.transcriptions.create(model="whisper-1", file=f)
        text = r.text
    except Exception as e:
        return {**_empty_result("audio:failed"), "missing": [str(e)]}
    return await _llm_extract(
        text=text, source_type="audio", filename=filename, parser_label="audio:transcript"
    )


# ── Public dispatcher ──────────────────────────────────────────────────────


async def ingest(input: dict[str, Any]) -> dict[str, Any]:
    src = input["source"]
    if src["kind"] == "text":
        return await parse_text(
            {"text": src["text"], "source_type": src.get("source_type", "user")}
        )

    buf = base64.b64decode(src["contents_b64"])
    mime = (src.get("mime") or "application/octet-stream").lower()
    filename = src["filename"]
    lower = filename.lower()

    if mime == "application/pdf" or lower.endswith(".pdf"):
        return await _parse_pdf(buf, filename)
    if "spreadsheetml" in mime or re.search(r"\.xlsx?$", lower):
        return await _parse_xlsx(buf, filename)
    if mime == "text/csv" or lower.endswith(".csv"):
        return await _parse_csv(buf, filename)
    if "wordprocessingml" in mime or lower.endswith(".docx"):
        return await _parse_docx(buf, filename)
    if mime.startswith("text/") or re.search(r"\.(md|markdown|txt)$", lower):
        return await parse_text(
            {"text": buf.decode("utf-8", errors="ignore"), "source_type": "md", "filename": filename}
        )
    if mime.startswith("image/"):
        return await _parse_image(buf, filename, mime)
    if mime.startswith("audio/") or mime.startswith("video/"):
        return await _parse_audio(buf, filename, mime)
    return await parse_text(
        {"text": buf.decode("utf-8", errors="ignore")[:100_000], "source_type": "md", "filename": filename}
    )
