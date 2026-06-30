"""Generate the pristine CFP master template shipped with the app.

Takes the firm's source workbook and clears the SAMPLE client's data from the
input cells (every non-formula cell on the INJECT tabs) while preserving all
formulas and all frozen-tab assumptions. The result is committed as
``master/cfp_master.xlsx`` and is what the engine injects fresh client data into.

Run once whenever the firm ships a new template:
    uv run python -m stackwealth.excel_engine.build_master <source.xlsx>
"""

from __future__ import annotations

import os
import sys

import openpyxl

from . import cellmap
from .engine import _is_formula

_HERE = os.path.dirname(__file__)
OUT_PATH = os.path.join(_HERE, "master", "cfp_master.xlsx")

# Header/label rows & columns to KEEP when clearing an input tab. Labels are
# identical between master and upload, so keeping them is harmless and makes the
# blank template human-readable. We clear only data-bearing cells: anything
# numeric, or text below the header band. Headers live in the first two rows or
# the first column of each tab in this template.
def _looks_like_label(coord_row: int, coord_col: int, value) -> bool:
    if not isinstance(value, str):
        return False
    # Keep any non-numeric TEXT label, wherever it sits — the firm's category
    # labels live in column E (and elsewhere), not just rows 1-2 / column A.
    # A string that parses as a number/currency is sample DATA, so clear it; a
    # plain text label ("Gross Salary per month", "EMI - Home Loan") is kept so
    # the blank template renders with labels on every path.
    from ..skills.intake_sheets import _amt

    if coord_row <= 2 or coord_col == 1:
        return True
    return _amt(value) is None


def build(source: str) -> str:
    wb = openpyxl.load_workbook(source, data_only=False)
    cleared = 0
    # Only clear the PURE input tabs. The compute/manual tabs (YoY, Retirement)
    # keep all their structure, anchors (e.g. the YoY year seed) and the firm's
    # RM-manual values, because the standard input upload doesn't re-supply them
    # — clearing them would compute the plan off zeros (broken year column,
    # missing one-time retirement spend, etc.).
    for tab in cellmap.INPUT_TABS:
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if v is None:
                    continue
                if _is_formula(v):
                    continue  # keep formulas (incl. array / data-table formulas)
                if _looks_like_label(c.row, c.column, v):
                    continue  # keep header / label text
                c.value = None
                cleared += 1
    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Cleared {cleared} sample-data cells across {len(cellmap.INPUT_TABS)} input tabs")
    print(f"Master written -> {OUT_PATH}")
    return OUT_PATH


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("usage: python -m stackwealth.excel_engine.build_master <source.xlsx>")
        raise SystemExit(2)
    build(sys.argv[1])
