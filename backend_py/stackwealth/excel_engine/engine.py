"""CFP Excel engine — the transformer.

Flow:
    client upload (.xlsx, firm template)
        -> inject its input cells into a pristine master copy
        -> recalc with LibreOffice
        -> read the result cells back
        -> (populated workbook bytes, structured outputs)

The populated workbook is what the UI's "Open computed Excel" button downloads;
the structured outputs feed the model / Python advisory layer.
"""

from __future__ import annotations

import io
import os
import shutil
import tempfile
from typing import Any

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula

from . import cellmap
from .recalc import recalc_file

_HERE = os.path.dirname(__file__)
MASTER_PATH = os.path.join(_HERE, "master", "cfp_master.xlsx")


def _is_formula(v: Any) -> bool:
    """True for an ordinary '='-prefixed formula string — these are preserved
    and recalculated.

    Array / data-table formulas (openpyxl objects) are deliberately NOT treated
    as formulas: openpyxl can't round-trip them cleanly and LibreOffice then
    recalcs them to #NAME? errors. They're cleared in the master and filled from
    the upload's cached value instead (correct for the standard input template,
    where e.g. 'years to go' is already evaluated)."""
    return isinstance(v, str) and v.startswith("=")


def _is_array_formula(v: Any) -> bool:
    return isinstance(v, (ArrayFormula, DataTableFormula))


def inject_inputs(master_wb, upload_wb_values, upload_wb_formulas) -> int:
    """Inject the upload's input cells into ``master_wb`` (mutated in place).

    Two policies, by tab type:

    • INPUT_TABS (1_Personal … 10_Goals, Risk) — MIRROR. Walk the master's range
      and set every non-formula cell to the upload's value at that coordinate,
      INCLUDING blanks, so a sparse upload clears the cleaned master's input
      slots and can never leak sample data. Real uploads always carry these tabs
      with their labels, so mirroring never wipes a header.

    • MANUAL_TABS (YoY Cash Flow, Retirement Plan) — OVERLAY, non-destructive.
      These hold the firm's structure, year anchor and RM-manual judgment cells
      and are NOT cleared in the master. A standard input upload doesn't contain
      them, so they're only touched when the upload actually carries the tab, and
      then only NON-EMPTY upload cells are written — the master's values are
      never wiped.

    Formula cells in the master are never overwritten. Returns cells written.
    """
    written = 0
    for tab in cellmap.INJECT_TABS:
        if tab not in master_wb.sheetnames:
            continue
        ms = master_wb[tab]
        uv = upload_wb_values[tab] if tab in upload_wb_values.sheetnames else None
        is_manual = tab in cellmap.MANUAL_TABS
        # Manual/compute tab absent from the upload → leave the master untouched.
        if is_manual and uv is None:
            continue
        max_row = ms.max_row
        max_col = ms.max_column
        if uv is not None:
            max_row = max(max_row, uv.max_row)
            max_col = max(max_col, uv.max_column)
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                mcell = ms.cell(row=r, column=c)
                if _is_formula(mcell.value):
                    continue  # never overwrite a firm formula
                up_val = uv.cell(row=r, column=c).value if uv is not None else None
                if is_manual and up_val is None:
                    continue  # overlay: never wipe a master cell with a blank
                if mcell.value != up_val:
                    mcell.value = up_val
                    written += 1
    return written


def _read_cell(ws, coord: str):
    try:
        return ws[coord].value
    except Exception:
        return None


def set_salary_horizon(master_wb, base_age: float | None, retire_age: float | None) -> None:
    """Re-shape the YoY salary column (E) to stop at the client's actual
    retirement age. The firm template hardcodes salary to grow for exactly 20
    rows (Y0–Y20, the sample client's working life) and is blank after — so
    without this every client's salary stops at ~age 56 regardless of their
    retirement age. Salary grows through the retirement year, then blanks;
    business / rental / other income columns are untouched."""
    if base_age is None or retire_age is None:
        return
    ws = master_wb["YoY Cash Flow"]
    FIRST, LAST = 6, 60
    retire_row = FIRST + max(0, round(retire_age - base_age))
    for r in range(FIRST + 1, LAST + 1):
        if r <= retire_row:
            cur = ws[f"E{r}"].value
            if not (isinstance(cur, str) and cur.startswith("=")):
                ws[f"E{r}"] = f"=E{r - 1}*(1+E$5)"
        else:
            ws[f"E{r}"] = None


_ASSET_GOAL_WORDS = ("house", "property", "flat", "real estate", "plot", "apartment", "villa")


def apply_house_to_nfa(master_wb) -> None:
    """Treat house / property purchase goals as ASSET ACQUISITIONS: convert the
    financial-asset outflow into a non-financial asset instead of consuming it.

    The firm template (and its own output) drains FA via the goal withdrawal and
    adds nothing to NFA, so a house purchase destroys net worth. Here we write
    the goal's future value into the YoY 'Addition of Fixed Assets' column (Y) in
    the purchase year — equal to what the same year withdraws from FA — so the
    purchase becomes NFA and net worth is preserved, then appreciates.

    Reads goal names from 10_Financial_Goals (populated by either injection or
    the model-writer), so it works for both paths. Only rows 3-9 are considered
    (the firm's SUMIF withdrawal range), keeping FA-out and NFA-in balanced."""
    g = master_wb["10_Financial_Goals"]
    yoy = master_wb["YoY Cash Flow"]
    base = yoy["C6"].value
    if not isinstance(base, (int, float)) or isinstance(base, bool):
        return
    base = int(base)
    additions: dict[int, list[str]] = {}
    for gr in range(3, 10):
        name = g[f"A{gr}"].value
        year = g[f"C{gr}"].value
        if not (isinstance(name, str) and isinstance(year, (int, float))):
            continue
        if not any(w in name.lower() for w in _ASSET_GOAL_WORDS):
            continue
        yrow = 6 + (int(year) - base)
        if 6 <= yrow <= 60:
            additions.setdefault(yrow, []).append(f"'10_Financial_Goals'!H{gr}")
    for yrow, refs in additions.items():
        existing = yoy[f"Y{yrow}"].value
        parts = list(refs)
        if isinstance(existing, (int, float)) and not isinstance(existing, bool) and existing:
            parts.insert(0, repr(existing))  # preserve a manual disposal/addition
        yoy[f"Y{yrow}"] = "=" + "+".join(parts)


def _age_retire_from_master(master_wb) -> tuple[float | None, float | None]:
    """Read base age (from DOB C5 + current-date D3) and retirement age (H5)
    out of an injected master's 1_Personal_Details tab."""
    from datetime import datetime

    ws = master_wb["1_Personal_Details"]
    dob = ws["C5"].value
    today = ws["D3"].value
    h5 = ws["H5"].value
    base_age = None
    if isinstance(dob, datetime) and isinstance(today, datetime):
        base_age = (today - dob).days / 365.25
    retire = float(h5) if isinstance(h5, (int, float)) and not isinstance(h5, bool) else None
    return base_age, retire


def extract_outputs(recalc_wb) -> dict[str, Any]:
    """Pull the headline scalars and result tables out of a recalculated book."""
    out: dict[str, Any] = {"scalars": {}, "tables": {}}

    for key, (sheet, coord) in cellmap.SCALAR_OUTPUTS.items():
        if sheet in recalc_wb.sheetnames:
            out["scalars"][key] = _read_cell(recalc_wb[sheet], coord)
        else:
            out["scalars"][key] = None

    for tname, spec in cellmap.TABLE_OUTPUTS.items():
        sheet = spec["sheet"]
        if sheet not in recalc_wb.sheetnames:
            out["tables"][tname] = []
            continue
        ws = recalc_wb[sheet]
        anchor = column_index_from_string(spec["anchor_col"])
        cols = {name: column_index_from_string(c) for name, c in spec["columns"].items()}
        rows = []
        r = spec["first_row"]
        blanks = 0
        while r <= ws.max_row and blanks < 3:
            if ws.cell(row=r, column=anchor).value in (None, ""):
                blanks += 1
                r += 1
                continue
            blanks = 0
            rows.append({name: ws.cell(row=r, column=ci).value for name, ci in cols.items()})
            r += 1
        out["tables"][tname] = rows
    return out


# Sheets surfaced in the in-app viewer, in display order. The computed result
# tabs first, then the key input tabs for reference.
VIEW_SHEETS: list[str] = [
    "10_Financial_Goals",
    "Retirement Plan",
    "YoY Cash Flow",
    "Insurance",
    "11. Inc Exp,Networth,Rec Invest",
    "2_Income",
    "3_Expenses ",
]


def _fmt_cell(value: Any, number_format: str | None) -> str:
    """Render a recalculated cell value to a display string, honouring the
    Excel number format loosely (percent / thousands / plain)."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        nf = number_format or ""
        if "%" in nf:
            return f"{value * 100:.2f}%"
        if "0" in nf and ("," in nf or "#,##" in nf):
            return f"{value:,.0f}" if abs(value) >= 100 else f"{value:,.2f}"
        # default: trim floats, keep ints clean
        if float(value).is_integer():
            return f"{int(value):,}"
        return f"{value:,.2f}"
    return str(value)


def _formula_text(v: Any) -> str | None:
    """Human-readable formula for a cell, or None for a plain value/label."""
    if isinstance(v, ArrayFormula):
        t = getattr(v, "text", None)
        return ("{%s}" % t) if t else None
    if isinstance(v, DataTableFormula):
        return "{table}"
    if isinstance(v, str) and v.startswith("="):
        return v
    return None


def render_sheets(recalc_bytes: bytes, sheets: list[str] | None = None) -> list[dict[str, Any]]:
    """Build an Excel-like grid representation of the recalculated workbook for
    the in-app viewer. Each cell carries its displayed VALUE and, when it's a
    formula, the FORMULA text — so the UI can show the calculation behind any
    cell (like Excel's formula bar). No desktop spreadsheet involved.

    Returns a list of:
        {name, cols: ["A","B",...], rows: [{"n": <row#>, "cells": [{v, f, num}]}]}
    """
    wb_v = openpyxl.load_workbook(io.BytesIO(recalc_bytes), data_only=True)
    wb_f = openpyxl.load_workbook(io.BytesIO(recalc_bytes), data_only=False)
    want = sheets or VIEW_SHEETS
    out: list[dict[str, Any]] = []
    for name in want:
        if name not in wb_v.sheetnames:
            continue
        wsv = wb_v[name]
        wsf = wb_f[name]
        # Trim to the used range (ignore trailing empties).
        last_row = last_col = 0
        for r in range(1, wsv.max_row + 1):
            for c in range(1, wsv.max_column + 1):
                if wsv.cell(r, c).value not in (None, ""):
                    last_row = max(last_row, r)
                    last_col = max(last_col, c)
        if last_row == 0:
            continue
        cols = [get_column_letter(c) for c in range(1, last_col + 1)]
        rows = []
        for r in range(1, last_row + 1):
            cells = []
            for c in range(1, last_col + 1):
                vcell = wsv.cell(r, c)
                disp = _fmt_cell(vcell.value, vcell.number_format)
                cells.append(
                    {
                        "v": disp,
                        "f": _formula_text(wsf.cell(r, c).value),
                        "num": isinstance(vcell.value, (int, float))
                        and not isinstance(vcell.value, bool),
                    }
                )
            rows.append({"n": r, "cells": cells})
        out.append({"name": name.strip(), "cols": cols, "rows": rows})
    return out


def compute_from_upload(
    upload_bytes: bytes,
    *,
    plan=None,
    master_path: str | None = None,
    timeout: int = 180,
) -> tuple[bytes, dict[str, Any]]:
    """Run the full pipeline on an uploaded firm-template workbook.

    When ``plan`` is given, the same DYNAMIC layer as the model-writer path is
    applied after injection (clear the sample's leaking lumpsums/remarks, allocate
    the client's real assets to goals per the firm rule, house->NFA, loan
    financing) — so firm-template uploads carry no hardcoded sample values either.
    Without a plan (e.g. the golden tests) it's pure injection.

    Returns (populated_recalculated_xlsx_bytes, structured_outputs).
    """
    master_path = master_path or MASTER_PATH
    if not os.path.exists(master_path):
        raise FileNotFoundError(
            f"CFP master template missing at {master_path}. Run build_master.py."
        )

    # Load the upload twice: formulas (to detect formula cells) and cached values.
    upload_formulas = openpyxl.load_workbook(io.BytesIO(upload_bytes), data_only=False)
    upload_values = openpyxl.load_workbook(io.BytesIO(upload_bytes), data_only=True)

    # Fresh master copy (formulas intact).
    master_wb = openpyxl.load_workbook(master_path, data_only=False)
    inject_inputs(master_wb, upload_values, upload_formulas)

    # Make salary respect the client's actual retirement age (a no-op when it
    # equals the template's frozen ~20-year horizon, e.g. retire-at-56).
    base_age, retire = _age_retire_from_master(master_wb)
    set_salary_horizon(master_wb, base_age, retire)

    if plan is not None:
        from .model_writer import (
            apply_dynamic_allocation,
            apply_loan_financing,
            apply_lumpsum_events,
            write_insurance_sheet,
        )

        apply_lumpsum_events(master_wb, plan)       # clear sample lumpsums/remarks
        apply_dynamic_allocation(master_wb, plan)   # categorise real assets → goals
        apply_house_to_nfa(master_wb)               # house -> NFA
        apply_loan_financing(master_wb, plan)       # down-payment + debt-netted NW
        write_insurance_sheet(master_wb, plan)      # platform-faithful Insurance tab
    else:
        apply_house_to_nfa(master_wb)

    return _recalc_and_extract(master_wb, timeout)


def compute_from_plan(
    plan, *, master_path: str | None = None, timeout: int = 180
) -> tuple[bytes, dict[str, Any]]:
    """Format-agnostic path: write a (LLM-extracted) PlanState into the master's
    input cells, recalculate, and read the results back. Used for uploads that
    aren't the firm's native template — the firm's formulas still do all the
    math, only the inputs are sourced from the normalised model.

    Returns (populated_recalculated_xlsx_bytes, structured_outputs).
    """
    from .model_writer import (
        apply_dynamic_allocation,
        apply_loan_financing,
        apply_lumpsum_events,
        write_insurance_sheet,
        write_plan_to_master,
    )

    master_path = master_path or MASTER_PATH
    if not os.path.exists(master_path):
        raise FileNotFoundError(
            f"CFP master template missing at {master_path}. Run build_master.py."
        )
    master_wb = openpyxl.load_workbook(master_path, data_only=False)
    write_plan_to_master(master_wb, plan)
    apply_lumpsum_events(master_wb, plan)      # clear sample lumpsums/remarks
    apply_dynamic_allocation(master_wb, plan)  # categorise real assets → goals
    apply_house_to_nfa(master_wb)
    apply_loan_financing(master_wb, plan)
    write_insurance_sheet(master_wb, plan)     # platform-faithful Insurance tab
    return _recalc_and_extract(master_wb, timeout)


def _mark_for_recalc(wb) -> None:
    """Force LibreOffice to recalculate every formula when it opens the workbook.

    openpyxl stamps `calcId="124519"` (a known calc-engine version) which makes
    LibreOffice/Excel trust the cached results and skip recalc. Since openpyxl
    can't compute formula caches, those cached values are stale/blank — so without
    this the recalc no-ops and every computed cell reads back as 0. calcId=0 +
    fullCalcOnLoad forces a full recalc on load. (Verified on LibreOffice 25.2.)"""
    wb.calculation.calcId = 0
    wb.calculation.fullCalcOnLoad = True


def _recalc_and_extract(master_wb, timeout: int) -> tuple[bytes, dict[str, Any]]:
    work = tempfile.mkdtemp(prefix="cfp_engine_")
    try:
        injected = os.path.join(work, "injected.xlsx")
        _mark_for_recalc(master_wb)
        master_wb.save(injected)

        recalced_path = recalc_file(injected, timeout=timeout)
        with open(recalced_path, "rb") as fh:
            populated_bytes = fh.read()

        recalc_wb = openpyxl.load_workbook(recalced_path, data_only=True)
        outputs = extract_outputs(recalc_wb)
        # clean up the recalc temp dir
        shutil.rmtree(os.path.dirname(recalced_path), ignore_errors=True)
        return populated_bytes, outputs
    finally:
        shutil.rmtree(work, ignore_errors=True)
