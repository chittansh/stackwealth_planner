"""Restore the firm's structural row/column LABELS into the master template.

An earlier `build_master.py` run stripped every text label that lives outside
rows 1-2 / column A (e.g. the category names in column E of `2_Income` and
`3_Expenses `), leaving the master's label columns empty. For firm-template
uploads the labels are mirrored back from the upload, but the model-writer path
(non-firm uploads) writes only values — so the computed-Excel pages render data
with no labels.

This one-time fixup copies a NON-FORMULA STRING from the firm source workbook
into the master at the SAME coordinate, but only when the master cell is empty
and not a formula. It is provably safe for the Excel math: on these tabs the
label columns feed no formula (totals are `=SUM(F:H)` over value columns only).

Run once (re-run is idempotent):
    uv run python -m stackwealth.excel_engine.restore_labels <firm_source.xlsx>

If no source is given it defaults to the firm template shipped in example_inputs.
"""

from __future__ import annotations

import os
import sys

import openpyxl

from . import cellmap
from .build_master import OUT_PATH
from .engine import _is_formula

_HERE = os.path.dirname(__file__)
# Default firm source: the template the master was built from (same tab names +
# row positions), still carrying every label.
_DEFAULT_SOURCE = os.path.abspath(
    os.path.join(_HERE, "..", "..", "..", "example_inputs",
                 "Format for inputs for CFP_ng_180626.xlsx")
)


def restore(source: str = _DEFAULT_SOURCE, master: str = OUT_PATH) -> int:
    src = openpyxl.load_workbook(source, data_only=False)
    mas = openpyxl.load_workbook(master, data_only=False)
    restored = 0
    for tab in cellmap.INPUT_TABS:
        if tab not in src.sheetnames or tab not in mas.sheetnames:
            continue
        sws, mws = src[tab], mas[tab]
        for row in sws.iter_rows():
            for c in row:
                v = c.value
                # Only copy plain TEXT labels — never numbers, never formulas.
                if not isinstance(v, str) or not v.strip() or _is_formula(v):
                    continue
                target = mws.cell(row=c.row, column=c.column)
                # Don't clobber an existing value or a formula in the master.
                if target.value is not None or _is_formula(target.value):
                    continue
                target.value = v
                restored += 1
    mas.save(master)
    print(f"Restored {restored} label cells across {len(cellmap.INPUT_TABS)} input tabs")
    print(f"Master updated -> {master}")
    return restored


if __name__ == "__main__":
    src = sys.argv[1] if len(sys.argv) > 1 else _DEFAULT_SOURCE
    if not os.path.exists(src):
        print(f"source not found: {src}")
        raise SystemExit(2)
    restore(src)
