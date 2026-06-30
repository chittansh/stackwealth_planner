"""Write a PlanState into the firm master template's input cells.

This is what makes the Excel engine *format-agnostic*. Direct cell-injection
(``engine.compute_from_upload``) only works when the upload is the firm's
standard template. For any other layout (e.g. the "Financial Planning_Client"
format with renamed tabs and a different cell structure), the LLM intake first
normalises it into a ``PlanState``; this module then writes that structured data
into the master's known input cells, and LibreOffice recalculates exactly as it
would for a native firm-template upload.

Only fields that drive the firm model are written. Missing fields are skipped
(the master's input cells were pre-cleared by build_master, so anything not
written stays blank).
"""

from __future__ import annotations

import re
from datetime import datetime
from typing import Any

# Base year the master's YoY anchor uses (the year seed on the YoY tab is kept
# from the template). Ages on 1_Personal_Details are computed as (D3 - DOB), so
# we anchor D3 to this base year for a consistent projection start.
BASE_YEAR = 2026
BASE_DATE = datetime(BASE_YEAR, 6, 1)


def _num(v: Any) -> float | None:
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    return None


def _to_date(v: Any) -> datetime | None:
    if isinstance(v, datetime):
        return v
    if isinstance(v, str) and v.strip():
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%m/%d/%Y"):
            try:
                return datetime.strptime(v.strip()[:10], fmt)
            except ValueError:
                continue
    return None


def _set(ws, coord: str, value: Any) -> None:
    if value is None:
        return
    ws[coord] = value


def write_assumption_overrides(master_wb, plan) -> int:
    """Push the plan's editable assumptions into the firm's 'Assumptions &
    Computation' tab so chat edits flow into the recalculated Excel.

    That tab is otherwise FROZEN (cellmap.FROZEN_TABS) — the firm's standard
    methodology — so without this a chat edit to income growth / inflation lands
    in the PlanState but the Excel keeps the master defaults (the bug behind
    "I changed it but the computed Excel still shows 5.6%"). We override only the
    cells whose plan default MATCHES the master, so an unedited plan is a no-op
    and only a real change moves the firm cell.

    Mapped today: per-source income growth + general inflation. (growth.* and
    taxes.* are baked into the tab's ROI/post-tax formulas — overriding them
    needs formula surgery and is left for a follow-up.)
    """
    a = getattr(plan, "assumptions", None)
    if a is None or "Assumptions & Computation" not in master_wb.sheetnames:
        return 0
    ws = master_wb["Assumptions & Computation"]
    written = 0

    # Income growth per source. The firm stores PRE-tax in col D and derives
    # POST-tax (col E = D*0.7) for the YoY projection (YoY!E5 = E12, etc.). The
    # PlanState stores POST-tax rates, so write the E cell directly (what the YoY
    # reads) and set D = E/0.7 for a consistent pre-tax display.
    ig = getattr(a, "income_growth", None)
    if ig is not None:
        for src, row in (("employment", 12), ("business", 13),
                         ("rental", 14), ("other", 17)):
            v = getattr(ig, src, None)
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                ws[f"E{row}"] = float(v)
                ws[f"D{row}"] = round(float(v) / 0.7, 4)
                written += 2

    # General household inflation (col D3). The per-type rows (education D4,
    # medical D5, …) aren't modelled per-client, so they keep the master values.
    infl = getattr(a, "inflation", None)
    if isinstance(infl, (int, float)) and not isinstance(infl, bool):
        ws["D3"] = float(infl)
        written += 1

    return written


# The firm template ships sample person names baked into header cells and
# descriptive labels (e.g. "Mr. M" / "Mrs. Y" column headers, "Life Expectancy of
# Mr M"). These are replaced with the ACTUAL client/spouse names from the input so
# the computed Excel never shows placeholder names. Client tokens vs spouse tokens
# are disjoint substrings; spouse (Mrs…) is substituted first.
_SPOUSE_NAME_TOKENS = ("Mrs. M", "Mrs M", "Mrs. Y", "Mrs Y")
_CLIENT_NAME_TOKENS = ("Mr. M", "Mr M", "Mr. X", "Mr X")
# Tabs the client actually sees where the sample names appear. Assumptions &
# Computation is frozen/internal and intentionally left untouched.
_NAME_TABS = (
    "1_Personal_Details", "2_Income", "3_Expenses ",
    "11. Inc Exp,Networth,Rec Invest", "Retirement Plan",
)


def _strip_age(name: str) -> str:
    """'Priya (35)' / 'Priya 32' / 'Mrs M - 40' → 'Priya' (drop a trailing age)."""
    if not name:
        return ""
    return re.split(r"[(\d]", str(name), 1)[0].strip(" -,").strip()


def write_person_names(master_wb, plan) -> int:
    """Replace the firm template's placeholder person names ('Mr. M' / 'Mrs. Y' …)
    with the real client and spouse names from the plan, across every cell on the
    client-facing tabs (headers AND descriptive labels). No-op without a plan or a
    client name. Returns the number of cells rewritten."""
    if plan is None:
        return 0
    pd = getattr(plan, "personal_details", None)
    persons = (plan.assumptions.persons if getattr(plan, "assumptions", None) else None) or []
    client = _strip_age(getattr(pd, "full_name", "") or "") or (
        _strip_age(getattr(persons[0], "name", "")) if persons else "")
    spouse = _strip_age(getattr(pd, "spouse_name_and_age", "") or "") or (
        _strip_age(getattr(persons[1], "name", "")) if len(persons) > 1 else "")
    if not client:
        return 0
    spouse = spouse or "Spouse"
    n = 0
    for tab in _NAME_TABS:
        if tab not in master_wb.sheetnames:
            continue
        ws = master_wb[tab]
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not isinstance(v, str) or v.startswith("="):
                    continue
                new = v
                for tok in _SPOUSE_NAME_TOKENS:
                    new = re.sub(re.escape(tok), spouse, new, flags=re.I)
                for tok in _CLIENT_NAME_TOKENS:
                    new = re.sub(re.escape(tok), client, new, flags=re.I)
                if new != v:
                    c.value = new
                    n += 1
    return n


def write_plan_to_master(master_wb, plan) -> int:
    """Populate ``master_wb`` (a fresh master copy) from ``plan``. Returns the
    number of cells written. Formula cells are left untouched."""
    written = 0

    def setc(ws, coord, value):
        nonlocal written
        if value is None:
            return
        ws[coord] = value
        written += 1

    # ── 1_Personal_Details ────────────────────────────────────────────────
    pd = plan.personal_details
    persons = (plan.assumptions.persons if plan.assumptions else None) or []
    p_self = persons[0] if len(persons) > 0 else None
    p_spouse = persons[1] if len(persons) > 1 else None
    ws = master_wb["1_Personal_Details"]
    setc(ws, "E2", BASE_YEAR)
    setc(ws, "D3", BASE_DATE)
    # Client
    dob = _to_date(getattr(pd, "date_of_birth", None)) or (
        _to_date(getattr(p_self, "date_of_birth", None)) if p_self else None
    )
    setc(ws, "C5", dob)
    retire = (
        _num(getattr(pd, "retirement_age_target", None))
        or (_num(getattr(p_self, "retirement_age", None)) if p_self else None)
        or 60
    )
    setc(ws, "H5", retire)
    setc(ws, "I5", (_num(getattr(p_self, "life_expectancy", None)) if p_self else None) or 85)
    # Spouse. The firm's retirement model derives the post-retirement horizon
    # from the SPOUSE's lifetime (provides until the dependent spouse passes).
    # For a single client, leaving the spouse cells blank yields a negative
    # horizon and a garbage corpus — so mirror the client into the spouse cells
    # (horizon = the client's own lifetime).
    spouse_life = _num(getattr(p_spouse, "life_expectancy", None)) if p_spouse else None
    spouse_dob = _to_date(getattr(p_spouse, "date_of_birth", None)) if p_spouse else None
    is_married = str(getattr(pd, "marital_status", "") or "").lower().startswith("married")
    if p_spouse and (spouse_dob or spouse_life) and is_married:
        setc(ws, "C6", spouse_dob or dob)
        setc(ws, "H6", _num(getattr(p_spouse, "retirement_age", None)) or retire)
        setc(ws, "I6", spouse_life or 85)
    else:
        # Single client → mirror self so the horizon is the client's lifetime.
        setc(ws, "C6", dob)
        setc(ws, "H6", retire)
        setc(ws, "I6", (_num(getattr(p_self, "life_expectancy", None)) if p_self else None) or 85)

    # ── 2_Income (rows 6-9 GROSS income; rows 17-18 deductions; F=client G=spouse)
    # Net income (I24) = gross − deductions, so when the source carries explicit
    # taxes / PF (e.g. the firm tab's Deductions section) they MUST be written or
    # the projection treats gross salary as take-home and overstates income.
    inc = plan.income_details
    if inc:
        ws = master_wb["2_Income"]
        setc(ws, "F6", _num(inc.client_salary_in_hand))
        setc(ws, "G6", _num(inc.spouse_salary_in_hand))
        setc(ws, "F7", _num(inc.client_business_income))
        setc(ws, "G7", _num(inc.spouse_business_income))
        setc(ws, "F8", _num(inc.client_rental_income))
        setc(ws, "G8", _num(inc.spouse_rental_income))
        setc(ws, "F9", _num(inc.client_other_income))
        setc(ws, "G9", _num(inc.spouse_other_income))
        setc(ws, "F17", _num(getattr(inc, "client_taxes", None)))      # Taxes from Salary
        setc(ws, "G17", _num(getattr(inc, "spouse_taxes", None)))
        setc(ws, "F18", _num(getattr(inc, "client_provident_fund", None)))  # Provident Fund
        setc(ws, "G18", _num(getattr(inc, "spouse_provident_fund", None)))

    # Salary in the YoY must stop at the client's actual retirement age, not the
    # template's frozen 20-year horizon.
    from .engine import set_salary_horizon

    base_age = (BASE_DATE - dob).days / 365.25 if dob else None
    set_salary_horizon(master_wb, base_age, retire)

    # The firm Retirement Plan reads "years to retire" for the SIP (PMT) from the
    # retirement goal rows (10_Financial_Goals!D19/D20). The model-writer has no
    # such RM-entered retirement goal, so those are 0 → PMT(rate, 0, …) = #NUM!.
    # Write the real years-to-retire there so the retirement SIP computes.
    if base_age is not None and retire and retire > base_age:
        yrs_to_retire = max(1, round(retire - base_age))
        gg = master_wb["10_Financial_Goals"]
        gg["D19"] = yrs_to_retire
        gg["D20"] = yrs_to_retire

        # Retirement horizon: the firm template reads life expectancy from the
        # Assumptions tab (self 75 / spouse 80) and bases the post-retirement
        # years on the SPOUSE — wrong for a single client, and ignores the
        # client's own life expectancy. Override E12/E13 with the client's
        # actual: for a single client the spouse mirrors self, so the horizon
        # (E15 = E13 − E14) becomes the client's own lifetime; for a couple the
        # corpus provides until the longer-living passes.
        rp = master_wb["Retirement Plan"]
        self_le = (_num(getattr(p_self, "life_expectancy", None)) if p_self else None) or 85
        rp["E12"] = self_le
        if p_spouse and is_married:
            spouse_le = _num(getattr(p_spouse, "life_expectancy", None)) or self_le
            rp["E13"] = max(self_le, spouse_le)
        else:
            rp["E13"] = self_le

    # ── 3_Expenses (value in col H, total I=SUM(F:H); map model fields to firm
    # rows). Every monthly_expenses field maps to its firm row so nothing is
    # dropped from the I26 total. Firm rows: 6 Rent, 7 Living, 8 Children, 9
    # Transport, 10 Utilities, 11 Other, 12 Lifestyle, 13 Medical, 14 Insurance
    # Health, 15 Insurance Life, 19 Entertainment/Discretionary.
    exp = plan.monthly_expenses
    if exp:
        ws = master_wb["3_Expenses "]
        setc(ws, "H6", _num(exp.rent_or_emi))
        setc(ws, "H7", (_num(exp.household_expenses) or 0) + (_num(exp.groceries) or 0) or None)
        setc(ws, "H8", _num(exp.school_fees))          # children → row 8 (post-retire excl.)
        setc(ws, "H9", _num(exp.transport))
        setc(ws, "H10", _num(exp.utilities))
        setc(ws, "H11", _num(exp.other_expenses))      # essential residual
        setc(ws, "H12", _num(exp.travel_or_lifestyle))
        setc(ws, "H13", _num(exp.medical))
        setc(ws, "H14", _num(exp.insurance_premium))   # insurance premium → Health row
        setc(ws, "H19", _num(exp.discretionary))       # discretionary → Entertainment row
        # NB: EMIs are NOT regular expenses — they go to the "Loan Repayments"
        # rows (23-25) below so they flow into the YoY loan-repayment column.

    # ── 5_Recurring_Investments (col C monthly amount, rows 2-7) ────────────
    mi = plan.monthly_investments
    if mi:
        ws = master_wb["5_Recurring_Investments"]
        setc(ws, "C2", _num(mi.mutual_fund_sip))
        setc(ws, "C3", _num(mi.nps))
        setc(ws, "C4", _num(mi.ppf))
        setc(ws, "C5", _num(mi.rd))
        setc(ws, "C6", _num(mi.direct_equity))
        setc(ws, "C12", _num(mi.insurance_premium))

    # Holding lists → firm sheets. last_row is the LAST data row before each
    # sheet's Total (verified against the master): MF 2-21, Equity 2-23, FI/RE/
    # Gold 2-9. If a client has MORE holdings than the template has rows, the
    # overflow's value is folded into the last row so the sheet Total never loses
    # money (the firm template caps individual rows; the portfolio total is what
    # drives the math).
    def _fill(ws, items, last_row, *, value_col, value_attr="current_value",
              name_col=None, name_attr=None, extra=()):
        items = list(items or [])
        cap = last_row - 1                      # rows 2..last_row inclusive
        head, tail = items[:cap], items[cap:]
        for i, h in enumerate(head):
            r = 2 + i
            if name_col and name_attr:
                setc(ws, f"{name_col}{r}", getattr(h, name_attr, None))
            setc(ws, f"{value_col}{r}", _num(getattr(h, value_attr, None)))
            for col, attr in extra:
                setc(ws, f"{col}{r}", _num(getattr(h, attr, None)))
        if tail and head:
            agg = (_num(getattr(head[-1], value_attr, None)) or 0) + sum(
                (_num(getattr(h, value_attr, None)) or 0) for h in tail)
            setc(ws, f"{value_col}{last_row}", agg)
            if name_col and name_attr:
                setc(ws, f"{name_col}{last_row}",
                     f"{getattr(head[-1], name_attr, None) or ''} +{len(tail)} more")

    _fill(master_wb["4A_Mutual_Funds"], plan.mutual_funds, 21,
          name_col="B", name_attr="fund_name", value_col="H",
          extra=[("I", "sip_amount")])
    _fill(master_wb["4B_Equity_Stocks"], plan.equity_stocks, 23,
          name_col="B", name_attr="stock_name", value_col="E")
    _fill(master_wb["4C_Fixed_Income"], plan.fixed_income, 9,
          name_col="C", name_attr="instrument", value_col="E",
          extra=[("D", "invested_amount")])
    _fill(master_wb["4D_Real_Estate"], plan.real_estate, 9, value_col="C")
    _fill(master_wb["4E_Gold & Others"], plan.gold, 9, value_col="C")

    # ── 6_Liquid_Capital (C amount) ────────────────────────────────────────
    lc = plan.liquid_capital
    if lc:
        ws = master_wb["6_Liquid_Capital"]
        setc(ws, "C2", _num(lc.savings_account_balance))
        setc(ws, "C3", _num(lc.idle_cash_for_investment))
        setc(ws, "C4", _num(lc.fd_breakable_for_investment))

    # ── 7_Emergency_Fund (C3 total corpus) ─────────────────────────────────
    ef = plan.emergency_fund
    if ef:
        ws = master_wb["7_Emergency_Fund"]
        setc(ws, "C3", _num(ef.total_emergency_corpus) or _num(ef.emergency_fund_available))

    # ── 8_Loans_Liabilities (C outstanding, D EMI, E rate) ─────────────────
    li = plan.loans_liabilities
    home_emi = 0.0
    other_structured_emi = 0.0
    if li:
        ws = master_wb["8_Loans_Liabilities"]
        rows = {"home_loan": 2, "car_loan": 3, "personal_loan": 4, "credit_card_dues": 5}
        for name, r in rows.items():
            blk = getattr(li, name, None)
            if blk is None:
                continue
            setc(ws, f"C{r}", _num(getattr(blk, "outstanding_amount", None)))
            setc(ws, f"D{r}", _num(getattr(blk, "emi", None)))
            setc(ws, f"E{r}", _num(getattr(blk, "interest_rate", None)))
            emi = _num(getattr(blk, "emi", None)) or 0.0
            if name == "home_loan":
                home_emi += emi
            else:
                other_structured_emi += emi

    # Loan EMIs feed the YoY 'Loan Repayments' column via 3_Expenses rows 23-25
    # (K = SUM(3_Expenses!I23:I25)*12). 8_Loans only drives the debt/insurance
    # views — it does NOT reach the cashflow — so EMIs must land here too.
    # Structured loans (loans_liabilities) are the source of truth; the generic
    # monthly_expenses.other_emis bucket is only used when there are no
    # structured loans, otherwise a chat that adds a loan AND bumps other_emis
    # would double-count.
    structured = home_emi + other_structured_emi
    other_emi = other_structured_emi
    if structured <= 0 and plan.monthly_expenses:
        other_emi = _num(getattr(plan.monthly_expenses, "other_emis", None)) or 0.0
    exp_ws = master_wb["3_Expenses "]
    if home_emi:
        setc(exp_ws, "H23", round(home_emi))     # EMI - Home Loan
    if other_emi:
        setc(exp_ws, "H24", round(other_emi))    # EMI - Other Loans

    # ── 9_Insurance_Details (E cover, F premium) ───────────────────────────
    ins = plan.insurance_details
    if ins:
        ws = master_wb["9_Insurance_Details"]
        term = getattr(ins, "term_plan", None)
        if term:
            setc(ws, "E2", _num(getattr(term, "cover_amount", None)))
            setc(ws, "F2", _num(getattr(term, "annual_premium", None)))
        health = getattr(ins, "health_insurance", None) or getattr(ins, "family_floater", None)
        if health:
            setc(ws, "E5", _num(getattr(health, "cover_amount", None)))
            setc(ws, "F5", _num(getattr(health, "annual_premium", None)))

    # ── 10_Financial_Goals (A name, C target year, E today's cost, F nature) ─
    # The master's E column is "Today's Cost" and the workbook inflates it to a
    # future value. So when the model only carries a FUTURE target (e.g. a file
    # whose goal column is "Future Value Needed"), discount it back to today's
    # money first — otherwise the workbook double-counts inflation.
    ws = master_wb["10_Financial_Goals"]
    NATURE = {"one_time": "One Time", "annual": "Annual", "recurring": "Annual"}
    for i, g in enumerate(plan.financial_goals or []):
        r = 3 + i
        if r > 17:
            break
        year = _num(getattr(g, "target_year", None))
        today = _num(getattr(g, "today_cost", None))
        target = _num(getattr(g, "target_amount", None))
        if today is None and target is not None:
            in_today = getattr(g, "is_target_in_today_money", True)
            if in_today is False and year and year > BASE_YEAR:
                infl = _num(getattr(g, "inflation_assumed", None)) or 0.07
                today = target / ((1 + infl) ** (int(year) - BASE_YEAR))
            else:
                today = target
        setc(ws, f"A{r}", getattr(g, "goal_name", None))
        setc(ws, f"C{r}", year)
        # Years-to-go (D) is an array formula in the firm template that we clear;
        # write it explicitly so the workbook inflates today's cost → future value.
        if year and year >= BASE_YEAR:
            setc(ws, f"D{r}", int(year) - BASE_YEAR)
        setc(ws, f"E{r}", round(today) if today is not None else None)
        freq = (getattr(g, "contribution_frequency", None) or "").lower()
        setc(ws, f"F{r}", NATURE.get(freq, "One Time"))

    # House/property purchases → NFA is applied centrally by
    # engine.apply_house_to_nfa after this writer runs (shared with the
    # firm-template path), reading the goal names just written above.
    return written


_HOUSE_WORDS = ("house", "property", "flat", "real estate", "plot", "apartment", "villa")
_LOAN_FIELDS = ("home_loan", "car_loan", "personal_loan", "credit_card_dues")
_PRIORITY_RANK = {"essential": 0, "important": 1, "aspirational": 2}
# Fixed-income instruments earmarked for retirement — reserved from goal funding.
_RETIREMENT_INSTRUMENTS = {"PPF", "EPF", "NPS", "SukanyaSamriddhi"}


def _holdings_value(rows) -> float:
    return sum((_num(getattr(r, "current_value", None)) or 0.0) for r in (rows or []))


def apply_lumpsum_events(master_wb, plan) -> None:
    """Clear the SAMPLE client's manual cashflow events leaking from the master's
    YoY and write the CLIENT's own. The firm sample hardcodes lumpsum deposits /
    property sales (col T), their remarks (col U — e.g. 'Knee Surgery planned…',
    'Reverse Mortgage', a 15Cr 'Balance sale') and fixed-asset disposals (col Y),
    none of which belong to other clients. Cleared here; house->NFA re-writes Y
    and the loan disbursement adds to T afterwards."""
    yws = master_wb["YoY Cash Flow"]
    for r in range(6, 57):
        for col in (20, 21, 25):  # T (lumpsum), U (remarks), Y (asset addition/disposal)
            cell = yws.cell(row=r, column=col)
            v = cell.value
            if v is not None and not (isinstance(v, str) and v.startswith("=")):
                cell.value = None
    events = (plan.assumptions.lumpsum_events if (plan and plan.assumptions) else None) or []
    for ev in events:
        yr = _num(getattr(ev, "year", None))
        amt = _num(getattr(ev, "amount", None))
        if yr and amt and yr >= BASE_YEAR:
            row = 6 + (int(yr) - BASE_YEAR)
            if 6 <= row <= 56:
                cur = yws.cell(row=row, column=20).value
                base_amt = cur if isinstance(cur, (int, float)) and not isinstance(cur, bool) else 0
                yws.cell(row=row, column=20).value = round(base_amt + amt)
                label = getattr(ev, "label", None)
                if label:
                    yws.cell(row=row, column=21).value = str(label)


def _asset_buckets(plan) -> list[list]:
    """Client's assets categorised in the firm's liquidation-priority order
    (mutable [label, amount] entries). Weak/Neutral stock tags aren't in the
    model, so trading stocks ≈ 'weak' (sold first) and long-term ≈ 'neutral'."""
    fi = plan.fixed_income or []
    fi_by: dict[str, float] = {}
    for r in fi:
        fi_by[getattr(r, "instrument", None) or "Other"] = fi_by.get(
            getattr(r, "instrument", None) or "Other", 0.0
        ) + (_num(getattr(r, "current_value", None)) or 0.0)
    stocks = plan.equity_stocks or []
    trading = sum(
        _num(getattr(s, "current_value", None)) or 0.0
        for s in stocks
        if (getattr(s, "long_term_or_trading", None) or "") == "trading"
    )
    longterm = sum(
        _num(getattr(s, "current_value", None)) or 0.0
        for s in stocks
        if (getattr(s, "long_term_or_trading", None) or "") != "trading"
    )
    liquid = 0.0
    lc = plan.liquid_capital
    if lc:
        liquid = (_num(getattr(lc, "idle_cash_for_investment", None)) or 0.0) + (
            _num(getattr(lc, "fd_breakable_for_investment", None)) or 0.0
        )
    re_sale = sum(
        _num(getattr(r, "current_value", None)) or 0.0
        for r in (plan.real_estate or [])
        if getattr(r, "earmarked_for_sale", False)
    )
    ordered = [
        ["Weak / Trading Stocks", trading],
        ["Liquid (idle / breakable)", liquid],
        ["Fixed Deposits", fi_by.get("FD", 0.0) + fi_by.get("RD", 0.0)],
        ["Bonds", fi_by.get("Bonds", 0.0)],
        ["Equity Stocks", longterm],
        ["Mutual Funds", _holdings_value(plan.mutual_funds)],
        ["NSC", fi_by.get("NSC", 0.0) + fi_by.get("PostOffice", 0.0)],
        ["PPF", fi_by.get("PPF", 0.0)],
        ["Real Estate (for sale)", re_sale],
        ["Gold", _holdings_value(plan.gold)],
        ["EPF", fi_by.get("EPF", 0.0)],
        ["NPS / Pension", fi_by.get("NPS", 0.0)],
    ]
    return [b for b in ordered if b[1] > 0]


def apply_dynamic_allocation(master_wb, plan) -> None:
    """Categorise the client's ACTUAL assets and allocate them to goals per the
    firm's documented rule, instead of inheriting the template's hardcoded sample
    allocation (e.g. Car ← '4B'!E26 + =839368).

    Firm rule (from the planning sheet):
      1. Sort financial needs chronologically (nearest goal first).
      2. Fund each goal by liquidating assets in this priority order: weak stocks,
         liquid, FDs, bonds, neutral stocks, MFs, NSC, PPF, real estate for sale,
         gold, EPF, pension — keeping the asset's exit just before the need.
      3. Never allocate more than the goal's need; the gap is covered by SIP
         (the workbook's PMT). Retirement is funded on its own tab → excluded.
    """
    g = master_wb["10_Financial_Goals"]
    yoy = master_wb["YoY Cash Flow"]
    c6 = yoy["C6"].value
    base_year = int(c6) if isinstance(c6, (int, float)) else BASE_YEAR

    # Clear the sample's asset→goal assignment (cols J..W); keep the goal math (X..).
    for r in range(3, 18):
        for col in range(10, 24):  # J(10) .. W(23)
            g.cell(row=r, column=col).value = None

    buckets = _asset_buckets(plan)

    # Goals chronological (nearest first); essential before desirable as a tiebreak.
    indexed = []
    for i, goal in enumerate(plan.financial_goals or []):
        row = 3 + i
        if row > 17:
            break
        if getattr(goal, "kind", "") == "retirement":
            continue
        yr = _num(getattr(goal, "target_year", None)) or (base_year + 99)
        pr = _PRIORITY_RANK.get((getattr(goal, "priority", None) or "important"), 1)
        need = _num(getattr(goal, "today_cost", None))
        if need is None:
            tgt = _num(getattr(goal, "target_amount", None)) or 0.0
            if getattr(goal, "is_target_in_today_money", True) is False and yr > base_year:
                infl = _num(getattr(goal, "inflation_assumed", None)) or 0.07
                need = tgt / ((1 + infl) ** (int(yr) - base_year))
            else:
                need = tgt
        indexed.append((int(yr), pr, row, need or 0.0))
    indexed.sort(key=lambda x: (x[0], x[1]))

    # Liquidate assets into each goal up to its need; record the per-asset
    # breakdown into the goal's asset slots (J/K, L/M, … up to 7 pairs).
    for _yr, _pr, row, need in indexed:
        remaining = need
        slot = 0
        for bucket in buckets:
            if remaining <= 0 or slot >= 7:
                break
            if bucket[1] <= 0:
                continue
            take = min(bucket[1], remaining)
            bucket[1] -= take
            remaining -= take
            name_col = 10 + slot * 2  # J, L, N, P, R, T, V
            g.cell(row=row, column=name_col).value = bucket[0]
            g.cell(row=row, column=name_col + 1).value = round(take)
            slot += 1

    # 5. Retirement step-up: the firm template hardcodes the starting annual
    #    contribution (E54 = 1,200,000) and rate (F51 = 10%). Drive them from the
    #    client's ACTUAL ongoing retirement investments (NPS + PPF + VPF, …) and
    #    their plan's step-up assumption.
    rp = master_wb["Retirement Plan"]
    mi = plan.monthly_investments
    ret_monthly = 0.0
    if mi:
        ret_monthly = (
            (_num(getattr(mi, "nps", None)) or 0.0)
            + (_num(getattr(mi, "ppf", None)) or 0.0)
            + (_num(getattr(mi, "other", None)) or 0.0)  # VPF / other retirement recurring
        )
    rp["E54"] = round(ret_monthly * 12) if ret_monthly > 0 else None
    step_up = _num(getattr(plan.assumptions, "sip_annual_step_up_pct", None)) if plan.assumptions else None
    if step_up is not None:
        # accept either a fraction (0.1) or a percent (10)
        rp["F51"] = step_up / 100.0 if step_up > 1 else step_up


def _loan_balance_series(specs, base_year, first_row, last_row):
    """Total outstanding loan balance per YoY row, amortised annually.
    specs: list of (outstanding, rate_pct, emi_monthly, start_year)."""
    out: dict[int, float] = {}
    for yi in range(0, last_row - first_row + 1):
        row, year = first_row + yi, base_year + yi
        total = 0.0
        for outstanding, rate, emi, start in specs:
            if year < start:
                continue
            bal = outstanding
            for _ in range(start, year):           # amortise from disbursement to this year
                interest = bal * (rate / 100.0)
                bal = max(0.0, bal - max(0.0, emi * 12 - interest))
                if bal <= 0:
                    break
            total += bal
        if total > 0:
            out[row] = round(total)
    return out


def apply_loan_financing(master_wb, plan) -> None:
    """Model loans as liabilities and (optionally) FINANCE a house purchase.

    1. Net worth net of debt: subtract each year's outstanding loan balance from
       the YoY net worth (the firm template never subtracts liabilities). Loan
       balances amortise toward zero as EMIs are paid.
    2. House funded by a loan: when there's a home loan AND a future house goal,
       treat the loan as financing that purchase — the loan disburses as cash in
       the purchase year (column T), so only the DOWN PAYMENT leaves financial
       assets; the house is still booked as NFA (apply_house_to_nfa) and the EMI
       services the debt. Net worth at purchase is therefore unchanged (cash →
       home equity), then grows as the loan amortises and the house appreciates.
    """
    li = plan.loans_liabilities
    if li is None:
        return
    yoy = master_wb["YoY Cash Flow"]
    c6 = yoy["C6"].value
    if not isinstance(c6, (int, float)):
        return
    base_year = int(c6)
    FIRST, LAST = 6, 56

    # Find the financed house goal (a future house/property purchase).
    house_goal = None
    for g in plan.financial_goals or []:
        nm = (getattr(g, "goal_name", "") or "").lower()
        yr = _num(getattr(g, "target_year", None))
        is_house = getattr(g, "kind", "") == "house_purchase" or any(w in nm for w in _HOUSE_WORDS)
        if is_house and yr and yr > base_year:
            house_goal = g
            break

    home = getattr(li, "home_loan", None)
    home_out = _num(getattr(home, "outstanding_amount", None)) if home else None
    fin_year = int(house_goal.target_year) if (house_goal and home_out) else None

    # Build amortisation specs (home loan starts at the purchase year when it
    # finances the house; all other loans are existing → start now).
    specs = []
    for name in _LOAN_FIELDS:
        blk = getattr(li, name, None)
        out = _num(getattr(blk, "outstanding_amount", None)) if blk else None
        if not out:
            continue
        rate = _num(getattr(blk, "interest_rate", None)) or 9.0
        emi = _num(getattr(blk, "emi", None)) or 0.0
        start = fin_year if (name == "home_loan" and fin_year) else base_year
        specs.append((out, rate, emi, start))
    if not specs:
        return

    # 2. Disbursement: add the financing loan as cash-in (T) in the purchase year.
    if fin_year and home_out:
        yrow = FIRST + (fin_year - base_year)
        if FIRST <= yrow <= LAST:
            cur = yoy[f"T{yrow}"].value
            expr = f"={round(home_out)}"
            if isinstance(cur, (int, float)) and not isinstance(cur, bool) and cur:
                expr += f"+{round(cur)}"
            yoy[f"T{yrow}"] = expr

    # 1. Net worth net of debt — write the loan balance (col AF) and subtract it.
    yoy["AF4"] = "Outstanding Loans"
    balances = _loan_balance_series(specs, base_year, FIRST, LAST)
    for row, bal in balances.items():
        yoy[f"AF{row}"] = bal
        ac = yoy[f"AC{row}"].value
        if isinstance(ac, str) and ac.startswith("="):
            yoy[f"AC{row}"] = f"=SUM(V{row},AA{row})-AF{row}"


def write_insurance_sheet(master_wb, plan) -> None:
    """Write a client-facing 'Insurance' sheet that mirrors the platform's
    insurance page — life-cover adequacy (avg of Human Life Value & needs-based
    corpus, plus loans, less existing cover and disposable assets) and health
    cover adequacy — using the SAME engine the page reads (compute_cfp's
    insurance block), so the computed-Excel tab and the page show identical
    numbers.

    Written as STATIC VALUES (not formulas): the rest of the workbook keeps its
    LibreOffice-recalculated formula caches, and a values-only sheet survives the
    recalc untouched. Non-fatal — any failure just skips the sheet."""
    try:
        from ..skills.cfp import compute_cfp

        ins = (compute_cfp(plan).insurance or {})
    except Exception:
        return
    if not ins:
        return
    from openpyxl.styles import Alignment, Font, PatternFill

    health = ins.get("health") or {}
    money = "#,##0"
    hdr_fill = PatternFill("solid", fgColor="2F4A3A")
    hdr_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14, color="2F4A3A")
    bold = Font(bold=True)
    accent = Font(bold=True, color="9A3412")

    if "Insurance" in master_wb.sheetnames:
        del master_wb["Insurance"]
    ws = master_wb.create_sheet("Insurance")
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 40

    r = 1

    def row(label, value=None, *, fmt=money, font=None, note="", fill=False, label_font=None):
        nonlocal r
        c_a = ws.cell(row=r, column=1, value=label)
        if label_font:
            c_a.font = label_font
        if value is not None:
            c_b = ws.cell(row=r, column=2, value=value)
            c_b.number_format = fmt
            if font:
                c_b.font = font
            c_b.alignment = Alignment(horizontal="right")
        if note:
            ws.cell(row=r, column=3, value=note)
        if fill:
            for col in (1, 2, 3):
                ws.cell(row=r, column=col).fill = hdr_fill
                ws.cell(row=r, column=col).font = hdr_font
        r += 1

    def gap():
        nonlocal r
        r += 1

    ws.cell(row=r, column=1, value="Insurance Needs Analysis").font = title_font
    r += 1
    ws.cell(row=r, column=1, value="Same methodology as the platform's Insurance page.").font = Font(
        italic=True, color="71717A"
    )
    r += 2

    # ── Life cover adequacy ──────────────────────────────────────────────
    hlv = ins.get("human_life_value", 0) or 0
    needs = ins.get("needs_based_corpus", 0) or 0
    avg = ins.get("average", 0) or 0
    total_need = ins.get("total_need_including_loans", 0) or 0
    loans = max(0, total_need - avg)
    existing = ins.get("existing_cover", 0) or 0
    assets = ins.get("investable_assets", 0) or 0
    additional = ins.get("additional_cover_required", 0) or 0
    covered = existing + assets
    life_pct = (covered / total_need) if total_need else 0

    row("LIFE COVER ADEQUACY", fill=True)
    row("Method A — Human Life Value (PV of future income to retirement)", hlv)
    row("Method B — Needs-based corpus (PV of dependants' expenses)", needs)
    row("Average of both methods", avg)
    row("Add: outstanding loans", loans)
    row("Total cover needed", total_need, font=bold, label_font=bold)
    row("Less: existing term cover", existing)
    row("Less: disposable financial assets credited", assets)
    row("Additional cover required", additional, font=accent, label_font=bold)
    row("Coverage of need (existing cover + assets)", round(life_pct, 4), fmt="0.0%")
    gap()

    # ── Health cover adequacy ────────────────────────────────────────────
    h_req = health.get("required", 0) or 0
    h_base = health.get("family_base", 0) or 0
    h_senior = health.get("senior_parent_cover", 0) or 0
    h_existing = health.get("existing_cover", 0) or 0
    h_additional = health.get("additional_cover_required", 0) or 0
    h_pct = (h_existing / h_req) if h_req else 0

    row("MEDICAL / HEALTH COVER ADEQUACY", fill=True)
    row("Required cover (higher of 50% income or family base + senior parents)", h_req, font=bold, label_font=bold)
    row("  — family base cover", h_base)
    if h_senior:
        row("  — separate senior-parent policies", h_senior)
    row("Existing cover (health + family floater)", h_existing)
    row("Additional cover required", h_additional, font=accent, label_font=bold)
    row("Coverage of need", round(h_pct, 4), fmt="0.0%")
    gap()

    # ── Existing policies ────────────────────────────────────────────────
    row("EXISTING POLICIES", fill=True)
    idet = plan.insurance_details

    def policy(label, b):
        nonlocal r
        ws.cell(row=r, column=1, value=label)
        if not b or not (getattr(b, "cover_amount", None) or getattr(b, "company", None)):
            ws.cell(row=r, column=3, value="—")
            r += 1
            return
        cover = getattr(b, "cover_amount", 0) or 0
        prem = getattr(b, "annual_premium", 0) or 0
        cb = ws.cell(row=r, column=2, value=cover)
        cb.number_format = money
        cb.alignment = Alignment(horizontal="right")
        comp = getattr(b, "company", None) or "—"
        ws.cell(row=r, column=3, value=f"{comp} · premium {prem:,.0f}/yr" if prem else str(comp))
        r += 1

    if idet:
        policy("Term plan (life)", idet.term_plan)
        policy("Health insurance", idet.health_insurance)
        policy("Family floater", idet.family_floater)
        policy("ULIP / Endowment", idet.ulip_or_endowment)
    gap()

    # ── How the life cover is calculated (formula trail) ─────────────────
    steps = ins.get("computation_trace") or []
    if steps:
        row("HOW THE LIFE COVER IS CALCULATED", fill=True)
        c = ws.cell(row=r, column=1, value="Step")
        c.font = bold
        ws.cell(row=r, column=2, value="Result").font = bold
        ws.cell(row=r, column=3, value="Formula").font = bold
        r += 1
        for st in steps:
            ws.cell(row=r, column=1, value=str(st.get("label", "")))
            val = st.get("value")
            if isinstance(val, (int, float)):
                cb = ws.cell(row=r, column=2, value=val)
                cb.number_format = money if (st.get("unit") != "%") else "0.0000"
                cb.alignment = Alignment(horizontal="right")
            ws.cell(row=r, column=3, value=str(st.get("formula", "")))
            r += 1
