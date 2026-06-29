"""Golden tests for the CFP Excel engine (stackwealth.excel_engine).

No pytest dependency — run directly:
    uv run python test_excel_engine.py

Checks, per firm sample workbook:
  1. FIDELITY OF INJECTION — the engine's outputs (built by pouring the client's
     INPUT cells into the pristine master and recalculating) must equal that
     client's OWN workbook recalculated directly. If they match, we reconstructed
     the plan exactly from inputs alone.
  2. MASTER ISOLATION — frozen tabs (Assumptions) and all master formulas are
     never mutated by injection; two different clients yield different outputs.
  3. OUTPUT SANITY — headline scalars are present and plausible.
"""
from __future__ import annotations

import io
import math
import os
import sys
from pathlib import Path

import openpyxl

from stackwealth.excel_engine import cellmap
from stackwealth.excel_engine.engine import compute_from_upload, MASTER_PATH
from stackwealth.excel_engine.recalc import recalc_file

EXAMPLES = Path(__file__).resolve().parent.parent / "example_inputs"
SAMPLES = [
    "Test profil_ng_180626.xlsx",
    "Naga_Inputs.xlsx",
    "V1_Young_Professional_Rohan.xlsx",
    "V4_Pre_Retirement_Rajeev.xlsx",
]


def _num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _close(a, b, rel=1e-4, abs_=1.0):
    if a is None and b is None:
        return True
    if not _num(a) or not _num(b):
        return str(a) == str(b)
    return math.isclose(a, b, rel_tol=rel, abs_tol=abs_)


FIRM_SOURCE = "Format for inputs for CFP_ng_180626.xlsx"


def _cached_outputs(path: str) -> dict:
    """Read the SAME output cells from a workbook's existing cached values
    (no recalc) — used to read the firm file's known-correct numbers."""
    from stackwealth.excel_engine.engine import extract_outputs

    wb = openpyxl.load_workbook(path, data_only=True)
    return extract_outputs(wb)


def test_roundtrip_fidelity():
    """The keystone golden test. The firm's own workbook has both sample INPUTS
    and known-correct cached OUTPUTS. Pour its inputs through the engine (clean
    master + inject + LibreOffice recalc) and the outputs must reproduce the
    firm file's cached numbers — proving the input->output path is exact.
    """
    print("\n== 1. ROUND-TRIP FIDELITY: engine(firm inputs) == firm cached outputs ==")
    path = str(EXAMPLES / FIRM_SOURCE)
    assert os.path.exists(path), f"firm source missing: {path}"
    _, engine_out = compute_from_upload(open(path, "rb").read())
    truth = _cached_outputs(path)

    mismatches = []
    for k, ev in engine_out["scalars"].items():
        # tax_* / debt_* are engine-INJECTED computations (model_calcs) the firm
        # file doesn't carry, so there's no cached counterpart to round-trip
        # against. Their correctness is checked separately against tax.py/debt.py.
        if k.startswith(("tax_", "debt_")):
            continue
        tv = truth["scalars"].get(k)
        if not _close(ev, tv):
            mismatches.append((k, ev, tv))
    eg, tg = engine_out["tables"]["goals"], truth["tables"]["goals"]
    if len(eg) != len(tg):
        mismatches.append(("goals_rowcount", len(eg), len(tg)))
    for i, (a, b) in enumerate(zip(eg, tg)):
        for col in ("future_value_needed", "required_sip", "gap_today"):
            if not _close(a.get(col), b.get(col)):
                mismatches.append((f"goal[{i}].{col}", a.get(col), b.get(col)))
    ey, ty = engine_out["tables"]["yoy_cashflow"], truth["tables"]["yoy_cashflow"]
    if len(ey) != len(ty):
        mismatches.append(("yoy_rowcount", len(ey), len(ty)))
    for i, (a, b) in enumerate(zip(ey, ty)):
        for col in ("income_employment", "expenses", "financial_assets_close"):
            if not _close(a.get(col), b.get(col), rel=1e-4, abs_=2.0):
                mismatches.append((f"yoy[{i}].{col}", a.get(col), b.get(col)))

    print(f"  scalars={len(engine_out['scalars'])} goals={len(eg)} yoy={len(ey)} "
          f"mismatches={len(mismatches)}")
    for m in mismatches[:10]:
        print(f"         mismatch {m}")
    assert not mismatches, "round-trip fidelity mismatch"
    print("  [OK ] engine reproduces the firm workbook's own numbers exactly")


def test_persona_sanity():
    """Engine produces plausible, non-null headline numbers on a CONFORMING
    firm-template file. Legacy-format files (V1-V4 / Naga, May-2016 layout) are
    printed for information only — they predate the 180626 template and are not
    guaranteed to map; real uploads use the current template."""
    print("\n== 4. PERSONA SANITY: plausible headline numbers ==")
    conforming = "Test profil_ng_180626.xlsx"
    path = str(EXAMPLES / conforming)
    _, out = compute_from_upload(open(path, "rb").read())
    s = out["scalars"]
    nw, corpus = s.get("net_worth"), s.get("retirement_corpus_required")
    ok = _num(nw) and nw > 0 and _num(corpus) and corpus > 0
    print(f"  [{'OK ' if ok else 'FAIL'}] {conforming:42} net_worth={nw:,} corpus={corpus:,.0f} "
          f"goals={len(out['tables']['goals'])}")
    assert ok, f"{conforming}: implausible headline numbers"

    for name in ["V1_Young_Professional_Rohan.xlsx", "V4_Pre_Retirement_Rajeev.xlsx"]:
        p = str(EXAMPLES / name)
        if not os.path.exists(p):
            continue
        _, o = compute_from_upload(open(p, "rb").read())
        print(f"  (info, legacy layout) {name:38} net_worth={o['scalars'].get('net_worth')} "
              f"corpus={o['scalars'].get('retirement_corpus_required')}")


def test_master_isolation():
    print("\n== 2. MASTER ISOLATION: formulas + frozen tabs never mutated ==")
    master = openpyxl.load_workbook(MASTER_PATH, data_only=False)
    # snapshot frozen-tab cells + a sample of formulas
    frozen_before = {}
    for tab in cellmap.FROZEN_TABS:
        if tab in master.sheetnames:
            ws = master[tab]
            frozen_before[tab] = {c.coordinate: c.value for row in ws.iter_rows() for c in row}
    formula_cells = []
    for tab in cellmap.INJECT_TABS:
        if tab in master.sheetnames:
            ws = master[tab]
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and c.value.startswith("="):
                        formula_cells.append((tab, c.coordinate, c.value))

    # run an injection (in-memory) and re-check the on-disk master is untouched
    path = str(EXAMPLES / SAMPLES[0])
    compute_from_upload(open(path, "rb").read())

    master2 = openpyxl.load_workbook(MASTER_PATH, data_only=False)
    ok = True
    for tab, cells in frozen_before.items():
        ws = master2[tab]
        for coord, val in cells.items():
            if ws[coord].value != val:
                ok = False
                print(f"  FAIL frozen {tab}!{coord} changed")
    for tab, coord, val in formula_cells[:500]:
        if master2[tab][coord].value != val:
            ok = False
            print(f"  FAIL formula {tab}!{coord} changed")
    assert ok, "master mutated by injection"
    print(f"  [OK ] {len(formula_cells)} formula cells + {len(cellmap.FROZEN_TABS)} frozen tabs intact "
          f"after injection (master is read-only)")


def test_cashflow_matches_format():
    """Regression guard for the year-anchor / compute-tab clearing bug. The YoY
    INCOME / EXPENSE / FINANCIAL-ASSET columns (A..V) must still tally cell-for
    -cell with the firm's Format reference. The NON-FINANCIAL-asset + net-worth
    columns (X..AC) now intentionally DIVERGE because house/property goals are
    converted FA->NFA (the firm treats them as consumption) — those are covered
    by test_house_to_nfa instead."""
    print("\n== 4. CASHFLOW FIDELITY: engine(Test-profil) YoY cols A..V == Format ==")
    populated, _ = compute_from_upload(open(str(EXAMPLES / "Test profil_ng_180626.xlsx"), "rb").read())
    eng = openpyxl.load_workbook(io.BytesIO(populated), data_only=True)["YoY Cash Flow"]
    ref = openpyxl.load_workbook(str(EXAMPLES / FIRM_SOURCE), data_only=True)["YoY Cash Flow"]
    LAST_CASHFLOW_COL = 22  # column V (Closing FA); NFA section starts at X (24)
    total, mism = 0, []
    for r in range(1, max(eng.max_row, ref.max_row) + 1):
        for c in range(1, LAST_CASHFLOW_COL + 1):
            a, b = eng.cell(r, c).value, ref.cell(r, c).value
            if _num(a) or _num(b):
                total += 1
                if not _close(a if _num(a) else 0, b if _num(b) else 0, abs_=0.5):
                    mism.append((eng.cell(r, c).coordinate, a, b))
    print(f"  YoY cashflow cells (A..V)={total} mismatches={len(mism)}")
    for m in mism[:8]:
        print(f"     {m}")
    assert not mism, "YoY cashflow (income/expense/FA) does not tally with the Format reference"
    print("  [OK ] cashflow (income/expense/FA) tallies exactly with the Excel format file")


def test_house_to_nfa():
    """House/property purchase goals must convert FA -> NFA, preserving net worth
    (vs the firm's model, which consumes them). Test-profil has a House Purchase
    in 2037; the engine's net worth that year must exceed the firm reference's by
    ~the house value, and the NFA closing balance must jump."""
    print("\n== 5. HOUSE -> NFA: house purchase preserves net worth ==")
    populated, _ = compute_from_upload(open(str(EXAMPLES / "Test profil_ng_180626.xlsx"), "rb").read())
    eng = openpyxl.load_workbook(io.BytesIO(populated), data_only=True)["YoY Cash Flow"]
    ref = openpyxl.load_workbook(str(EXAMPLES / FIRM_SOURCE), data_only=True)["YoY Cash Flow"]
    # find the 2037 row (House Purchase year), col Y=25 addition, AC=29 net worth
    def row_for_year(ws, year):
        for r in range(6, 60):
            if ws.cell(r, 3).value == year:
                return r
        return None
    r = row_for_year(eng, 2037)
    assert r, "2037 row not found"
    y_add = eng.cell(r, 25).value or 0
    nw_eng = eng.cell(r, 29).value or 0
    nw_ref = ref.cell(r, 29).value or 0
    print(f"  2037: NFA addition(Y)={y_add:,.0f} | net worth engine={nw_eng:,.0f} ref={nw_ref:,.0f} (Δ={nw_eng - nw_ref:,.0f})")
    assert y_add > 1_000_000, "house not added to NFA in 2037"
    assert nw_eng > nw_ref + 1_000_000, "house→NFA did not lift net worth vs firm reference"
    print("  [OK ] house purchase converted FA→NFA; net worth preserved")


def test_house_financing():
    """A house funded by a home loan: the loan disburses as cash in the purchase
    year (column T) so only the down-payment leaves FA, and the outstanding loan
    balance is subtracted from net worth (column AF; AC formula adjusted)."""
    print("\n== 6. HOUSE FINANCING: loan disbursement + net-worth net of debt ==")
    from types import SimpleNamespace
    from stackwealth.excel_engine.model_writer import apply_loan_financing
    from stackwealth.excel_engine.engine import MASTER_PATH

    plan = SimpleNamespace(
        financial_goals=[SimpleNamespace(goal_name="House Purchase", kind="house_purchase", target_year=2030)],
        loans_liabilities=SimpleNamespace(
            home_loan=SimpleNamespace(outstanding_amount=8_000_000, emi=70_000, interest_rate=8.5),
            car_loan=None, personal_loan=None, credit_card_dues=None,
        ),
        monthly_expenses=None,
    )
    wb = openpyxl.load_workbook(MASTER_PATH, data_only=False)
    yoy = wb["YoY Cash Flow"]
    yrow = 6 + (2030 - 2026)  # row 10
    apply_loan_financing(wb, plan)
    t = yoy[f"T{yrow}"].value
    af = yoy[f"AF{yrow}"].value
    ac = yoy[f"AC{yrow}"].value
    print(f"  purchase row {yrow}: T(disbursement)={t} | AF(loan bal)={af} | AC(net worth)={ac}")
    assert isinstance(t, str) and "8000000" in t, "loan not disbursed into T in purchase year"
    assert isinstance(af, (int, float)) and af > 7_000_000, "loan balance not booked"
    assert isinstance(ac, str) and "AF" in ac, "net worth not adjusted for the loan"
    print("  [OK ] house financed by loan: down-payment-only FA hit, loan netted off net worth")


def test_dynamic_allocation():
    """Goal asset allocation must come from the CLIENT's real holdings, not the
    firm template's hardcoded sample (e.g. =839368, or refs to specific sample
    asset cells). After allocation, the goal's 'Current Value' (K) equals the
    investable assets waterfalled to it, and the sample hardcode M3 is gone."""
    print("\n== 7. DYNAMIC ALLOCATION: goals funded from the client's real assets ==")
    from types import SimpleNamespace
    from stackwealth.excel_engine.model_writer import apply_dynamic_allocation
    from stackwealth.excel_engine.engine import MASTER_PATH

    plan = SimpleNamespace(
        equity_stocks=[SimpleNamespace(current_value=1_000_000)],
        mutual_funds=[SimpleNamespace(current_value=500_000)],
        fixed_income=[SimpleNamespace(current_value=300_000, instrument="FD")],
        gold=[], real_estate=[], liquid_capital=None, monthly_investments=None, assumptions=None,
        financial_goals=[
            SimpleNamespace(goal_name="Car", kind="other", target_year=2032,
                            today_cost=2_000_000, priority="essential",
                            target_amount=None, is_target_in_today_money=True,
                            inflation_assumed=None),
        ],
    )
    wb = openpyxl.load_workbook(MASTER_PATH, data_only=False)
    g = wb["10_Financial_Goals"]
    assert g["M3"].value == "=839368", "expected the sample hardcode in the master"
    apply_dynamic_allocation(wb, plan)
    # value cells: K,M,O,Q,S,U,W (cols 11,13,15,17,19,21,23)
    vals = {}
    for col in (11, 13, 15, 17, 19, 21, 23):
        v = g.cell(row=3, column=col).value
        if isinstance(v, (int, float)):
            vals[g.cell(row=3, column=col - 1).value] = v
    total = sum(vals.values())
    print(f"  allocation breakdown: {vals} (total={total})")
    # 0.3M FD + 1.0M equity + 0.5M MF = 1.8M of the client's REAL assets
    assert total == 1_800_000, f"expected 1.8M client assets allocated, got {total}"
    assert "Fixed Deposits" in vals and "Equity Stocks" in vals, "firm liquidation order not applied"
    assert g["M3"].value != "=839368", "sample hardcode =839368 not cleared"
    print("  [OK ] goal funded from the client's real assets in firm priority order; hardcode gone")


def test_two_clients_differ():
    print("\n== 3. DISTINCTNESS: different clients -> different net worth ==")
    nws = {}
    for name in SAMPLES[:2]:
        path = str(EXAMPLES / name)
        if not os.path.exists(path):
            continue
        _, out = compute_from_upload(open(path, "rb").read())
        nws[name] = out["scalars"].get("net_worth")
        print(f"  {name:42} net_worth={nws[name]}")
    vals = [v for v in nws.values() if _num(v)]
    assert len(set(vals)) == len(vals), "distinct clients produced identical net worth"
    print("  [OK ] outputs are client-specific")


if __name__ == "__main__":
    if not os.path.exists(MASTER_PATH):
        print(f"MASTER missing: {MASTER_PATH} — run build_master first")
        sys.exit(2)
    test_roundtrip_fidelity()
    test_master_isolation()
    test_cashflow_matches_format()
    test_house_to_nfa()
    test_house_financing()
    test_dynamic_allocation()
    test_two_clients_differ()
    test_persona_sanity()
    print("\nALL EXCEL-ENGINE GOLDEN TESTS PASSED")
