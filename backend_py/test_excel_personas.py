"""Run the Excel-faithful CFP engine on the 4 persona workbooks the
firm uses to QA the planner, and print a side-by-side comparison of
engine output vs the FV figures the Excel itself computed."""
from __future__ import annotations

import sys
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import load_workbook

from stackwealth.types import (
    PlanState, PersonalDetails, IncomeDetails, MonthlyExpenses,
    MFHolding, StockHolding, FixedIncomeRow, RealEstateHolding,
    GoldHolding, MonthlyInvestments, LiquidCapital, EmergencyFund,
    LoanBlock, Liabilities, InsuranceBlock, InsuranceDetails,
    Goal, Person, Assumptions, Growth, IncomeGrowth,
    FreedomScoreInputs, empty_plan_state,
)
from stackwealth.skills.cfp import compute_cfp
from stackwealth.skills.debt import compute_debt_ratios, compute_repayment_strategies
from stackwealth.skills.tax import compute_tax_regime_comparison
from stackwealth.skills.risk import compute_questionnaire_score, RISK_QUESTIONNAIRE
from stackwealth.skills.cashflow import compute_cashflow


GOAL_KIND_MAP = {
    "Child Education": "child_education",
    "Child Marriage": "child_marriage",
    "Retirement": "retirement",
    "House Purchase": "house_purchase",
    "Foreign Travel": "foreign_travel",
}


def _f(v: Any) -> float:
    if v is None or v == "":
        return 0
    if isinstance(v, str):
        s = v.strip().rstrip("%").replace(",", "")
        try:
            return float(s)
        except ValueError:
            return 0
    return float(v)


def _dob_to_iso(dob_str: str) -> str:
    """Excel format '15-Aug-1997' → DD-MM-YYYY string."""
    import datetime as dt
    months = {"Jan":1,"Feb":2,"Mar":3,"Apr":4,"May":5,"Jun":6,
              "Jul":7,"Aug":8,"Sep":9,"Oct":10,"Nov":11,"Dec":12}
    if isinstance(dob_str, dt.datetime):
        return dob_str.strftime("%d-%m-%Y")
    parts = str(dob_str).split("-")
    if len(parts) == 3 and parts[1] in months:
        d, m, y = parts
        return f"{int(d):02d}-{months[m]:02d}-{y}"
    return str(dob_str)


def build_plan_from_excel(path: Path) -> PlanState:
    wb = load_workbook(path, data_only=True)
    plan = empty_plan_state(f"persona:{path.stem}")

    # ── 1. Personal Details ────────────────────────────────────────────
    ws = wb["1_Personal_Details"]
    pd_map = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(2, ws.max_row + 1)}
    name = pd_map.get("Full Name") or ""
    dob = pd_map.get("Date of Birth")
    plan.personal_details = PersonalDetails(
        full_name=name,
        date_of_birth=_dob_to_iso(dob) if dob else None,
        marital_status=pd_map.get("Marital Status"),
        spouse_name_and_age=pd_map.get("Spouse Name & Age"),
        number_of_children=int(pd_map.get("Number of Children") or 0),
        dependents=int(pd_map.get("Dependents (Parents/Others)") and 1 or 0)
            if isinstance(pd_map.get("Dependents (Parents/Others)"), str) else 0,
        city_of_residence=pd_map.get("City of Residence"),
        city_type="Metro" if pd_map.get("City of Residence") in ("Bangalore","Mumbai","New Delhi","Pune","Chennai","Kolkata","Hyderabad") else "Non-metro",
        occupation=pd_map.get("Occupation"),
        retirement_age_target=int(pd_map.get("Retirement Age Target") or 60),
    )

    # Person record (for `assumptions.persons`)
    plan.assumptions.persons = [Person(
        id="primary",
        name=name,
        date_of_birth=_dob_to_iso(dob) if dob else None,
        life_expectancy=85,
        retirement_age=plan.personal_details.retirement_age_target,
    )]
    # Spouse if present
    spouse_str = pd_map.get("Spouse Name & Age") or ""
    if isinstance(spouse_str, str) and "," in spouse_str:
        spouse_name, spouse_age_str = (s.strip() for s in spouse_str.split(",", 1))
        try:
            spouse_age = int(spouse_age_str.split()[0])
            from datetime import datetime as _dt
            spouse_dob_yr = _dt.now().year - spouse_age
            plan.assumptions.persons.append(Person(
                id="spouse",
                name=spouse_name,
                date_of_birth=f"01-01-{spouse_dob_yr}",
                life_expectancy=85,
                retirement_age=plan.personal_details.retirement_age_target,
            ))
        except (ValueError, IndexError):
            pass

    # ── 2. Income Details ──────────────────────────────────────────────
    ws = wb["2_Income_Details"]
    inc = {ws.cell(r, 1).value: (ws.cell(r, 2).value, ws.cell(r, 3).value) for r in range(2, ws.max_row + 1)}
    plan.income_details = IncomeDetails(
        client_salary_in_hand=_f(inc.get("Salary (In-hand)", (0, 0))[0]),
        spouse_salary_in_hand=_f(inc.get("Salary (In-hand)", (0, 0))[1]),
        client_business_income=_f(inc.get("Business Income", (0, 0))[0]),
        spouse_business_income=_f(inc.get("Business Income", (0, 0))[1]),
        client_rental_income=_f(inc.get("Rental Income", (0, 0))[0]),
        spouse_rental_income=_f(inc.get("Rental Income", (0, 0))[1]),
        client_other_income=_f(inc.get("Other Income", (0, 0))[0]),
        spouse_other_income=_f(inc.get("Other Income", (0, 0))[1]),
    )
    monthly_income = sum([
        plan.income_details.client_salary_in_hand or 0,
        plan.income_details.spouse_salary_in_hand or 0,
        plan.income_details.client_business_income or 0,
        plan.income_details.spouse_business_income or 0,
        plan.income_details.client_rental_income or 0,
        plan.income_details.spouse_rental_income or 0,
        plan.income_details.client_other_income or 0,
        plan.income_details.spouse_other_income or 0,
    ])

    # ── 3. Monthly Expenses ────────────────────────────────────────────
    ws = wb["3_Monthly_Expenses"]
    me_map = {ws.cell(r, 1).value: _f(ws.cell(r, 2).value) for r in range(2, ws.max_row + 1)}
    plan.monthly_expenses = MonthlyExpenses(
        household_expenses=me_map.get("Household Expenses"),
        rent_or_emi=me_map.get("Rent / EMI"),
        groceries=me_map.get("Groceries"),
        utilities=me_map.get("Utilities"),
        school_fees=me_map.get("School Fees"),
        insurance_premium=me_map.get("Insurance Premium"),
        medical=me_map.get("Medical"),
        travel_or_lifestyle=me_map.get("Travel / Lifestyle"),
        sip_investments=me_map.get("SIP Investments"),
        other_emis=me_map.get("Other EMIs"),
    )

    # ── 4A. Mutual Funds ───────────────────────────────────────────────
    ws = wb["4A_Mutual_Funds"]
    for r in range(2, ws.max_row + 1):
        name_ = ws.cell(r, 1).value
        if not name_ or str(name_).lower() == "total":
            continue
        plan.mutual_funds.append(MFHolding(
            id=str(uuid4()),
            fund_name=str(name_),
            folio=str(ws.cell(r, 2).value or ""),
            current_value=_f(ws.cell(r, 3).value),
            sip_amount=_f(ws.cell(r, 4).value),
        ))

    # ── 4B. Stocks ─────────────────────────────────────────────────────
    ws = wb["4B_Equity_Stocks"]
    for r in range(2, ws.max_row + 1):
        name_ = ws.cell(r, 1).value
        if not name_ or str(name_).lower() == "total":
            continue
        long_or_trading = (ws.cell(r, 4).value or "").lower()
        plan.equity_stocks.append(StockHolding(
            id=str(uuid4()),
            stock_name=str(name_),
            quantity=_f(ws.cell(r, 2).value),
            current_value=_f(ws.cell(r, 3).value),
            long_term_or_trading="trading" if "trad" in long_or_trading else "long_term",
        ))

    # ── 4C. Fixed Income ───────────────────────────────────────────────
    ws = wb["4C_Fixed_Income"]
    for r in range(2, ws.max_row + 1):
        name_ = ws.cell(r, 1).value
        if not name_ or str(name_).lower() == "total":
            continue
        cv = _f(ws.cell(r, 3).value)
        if cv <= 0:
            continue
        # Excel labels: FD, RD, PPF, EPF, Bonds — match the Instrument literal.
        inst_label = str(name_).strip().upper()
        if inst_label not in ("FD", "RD", "PPF", "EPF", "BONDS", "NPS"):
            continue
        instr = "Bonds" if inst_label == "BONDS" else inst_label
        plan.fixed_income.append(FixedIncomeRow(
            id=str(uuid4()),
            instrument=instr,
            invested_amount=_f(ws.cell(r, 2).value),
            current_value=cv,
        ))

    # ── 4D. Real Estate ────────────────────────────────────────────────
    ws = wb["4D_Real_Estate"]
    for r in range(2, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if not label or "no real" in str(label).lower():
            continue
        plan.real_estate.append(RealEstateHolding(
            label=str(label),
            kind="residential",
            current_value=_f(ws.cell(r, 2).value),
            earmarked_for_sale=False,
        ))

    # ── 4E. Gold / Others ──────────────────────────────────────────────
    ws = wb["4E_Gold_Others"]
    for r in range(2, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if not label or str(label).lower() == "total":
            continue
        lower = str(label).lower()
        kind = "physical"
        if "sgb" in lower or "gold bond" in lower or "sovereign" in lower:
            kind = "sgb"
        elif "digital" in lower:
            kind = "digital"
        elif "jewell" in lower:
            kind = "jewellery"
        plan.gold.append(GoldHolding(
            label=str(label),
            kind=kind,
            current_value=_f(ws.cell(r, 2).value),
            held_for_investment=True,
        ))

    # ── 5. Monthly Investments ─────────────────────────────────────────
    ws = wb["5_Monthly_Investments"]
    mi_map = {ws.cell(r, 1).value: _f(ws.cell(r, 2).value) for r in range(2, ws.max_row + 1)}
    plan.monthly_investments = MonthlyInvestments(
        mutual_fund_sip=mi_map.get("Mutual Fund SIP"),
        nps=mi_map.get("NPS"),
        ppf=mi_map.get("PPF"),
        rd=mi_map.get("RD"),
        direct_equity=mi_map.get("Direct Equity"),
        insurance_premium=mi_map.get("Insurance Premium"),
        other=mi_map.get("Other"),
    )

    # ── 6. Liquid Capital ──────────────────────────────────────────────
    ws = wb["6_Liquid_Capital"]
    lc_map = {ws.cell(r, 1).value: _f(ws.cell(r, 2).value) for r in range(2, ws.max_row + 1)}
    plan.liquid_capital = LiquidCapital(
        savings_account_balance=lc_map.get("Savings Account Balance"),
        idle_cash_for_investment=lc_map.get("Idle Cash for Investment"),
        fd_breakable_for_investment=lc_map.get("FD Breakable for Investment"),
        bonus_expected_for_investment=lc_map.get("Bonus Expected for Investment"),
    )

    # ── 7. Emergency Fund ──────────────────────────────────────────────
    ws = wb["7_Emergency_Fund"]
    ef_map = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(2, ws.max_row + 1)}
    plan.emergency_fund = EmergencyFund(
        emergency_fund_available=str(ef_map.get("Emergency fund available? (Yes/No)") or "").lower() == "yes",
        total_emergency_corpus=_f(ef_map.get("Total Emergency Corpus")),
        where_is_it_parked=ef_map.get("Where is it parked?"),
        monthly_household_expense_for_calculation=_f(ef_map.get("Monthly household expense (for calculation)")),
        months_of_cover_available=_f(ef_map.get("Months of cover available")),
    )

    # ── 8. Loans ───────────────────────────────────────────────────────
    ws = wb["8_Loans_Liabilities"]
    loan_keys = {
        "Home Loan": "home_loan",
        "Car Loan": "car_loan",
        "Personal Loan": "personal_loan",
        "Credit Card Dues": "credit_card_dues",
    }
    liab_kwargs = {}
    for r in range(2, ws.max_row + 1):
        key = ws.cell(r, 1).value
        if key not in loan_keys:
            continue
        outstanding = _f(ws.cell(r, 2).value)
        if outstanding <= 0:
            continue
        rate_raw = ws.cell(r, 4).value
        rate = None
        if rate_raw:
            m = "".join(c for c in str(rate_raw) if c.isdigit() or c == ".")
            try:
                rate = float(m) if m else None
            except ValueError:
                rate = None
        tenure_raw = ws.cell(r, 5).value
        tenure = None
        if tenure_raw and "pay in full" not in str(tenure_raw).lower():
            m = "".join(c for c in str(tenure_raw) if c.isdigit() or c == ".")
            try:
                tenure = float(m) if m else None
            except ValueError:
                tenure = None
        liab_kwargs[loan_keys[key]] = LoanBlock(
            outstanding_amount=outstanding,
            emi=_f(ws.cell(r, 3).value),
            interest_rate=rate,
            tenure_left=tenure,
        )
    plan.loans_liabilities = Liabilities(**liab_kwargs)

    # ── 9. Insurance ───────────────────────────────────────────────────
    ws = wb["9_Insurance_Details"]
    ins_keys = {
        "Term Plan": "term_plan",
        "Health Insurance": "health_insurance",
        "Family Floater": "family_floater",
        "ULIP / Endowment": "ulip_or_endowment",
    }
    ins_kwargs = {}
    for r in range(2, ws.max_row + 1):
        key = ws.cell(r, 1).value
        if key not in ins_keys:
            continue
        cover = _f(ws.cell(r, 3).value)
        if cover <= 0 and not ws.cell(r, 2).value:
            continue
        ins_kwargs[ins_keys[key]] = InsuranceBlock(
            company=ws.cell(r, 2).value,
            cover_amount=cover,
            annual_premium=_f(ws.cell(r, 4).value),
        )
    plan.insurance_details = InsuranceDetails(**ins_kwargs)

    # ── 10. Goals ──────────────────────────────────────────────────────
    ws = wb["10_Financial_Goals"]
    for r in range(2, ws.max_row + 1):
        goal_label = ws.cell(r, 1).value
        if not goal_label or "total" in str(goal_label).lower():
            continue
        target_year = ws.cell(r, 2).value
        if not target_year:
            continue
        today_cost = _f(ws.cell(r, 3).value)
        if today_cost <= 0:
            continue
        infl_raw = ws.cell(r, 4).value
        infl = _f(infl_raw) / 100 if infl_raw else None
        kind = GOAL_KIND_MAP.get(str(goal_label), "other")
        plan.financial_goals.append(Goal(
            goal_name=str(goal_label),
            kind=kind,
            target_year=int(target_year),
            target_amount=today_cost,
            is_target_in_today_money=True,
            inflation_assumed=infl,
            horizon_years=int(target_year) - 2026,
            priority="essential" if kind in ("retirement","child_education","child_marriage") else "important",
        ))

    # ── FSI aggregates ─────────────────────────────────────────────────
    expense_keys_in_fsi = ("household_expenses","rent_or_emi","groceries","utilities",
                           "school_fees","insurance_premium","medical","travel_or_lifestyle")
    me_total = sum((getattr(plan.monthly_expenses, k) or 0) for k in expense_keys_in_fsi)
    emi_other = plan.monthly_expenses.other_emis or 0
    # Also include EMI from the loan block (e.g. credit card)
    loan_emi = 0
    for k in ("home_loan","car_loan","personal_loan","credit_card_dues"):
        b = getattr(plan.loans_liabilities, k, None)
        if b and b.emi:
            loan_emi += b.emi or 0
    portfolio_total = (sum(h.current_value or 0 for h in plan.mutual_funds)
                       + sum(h.current_value or 0 for h in plan.equity_stocks)
                       + sum(h.current_value or 0 for h in plan.fixed_income))
    liquid_total = sum([
        plan.liquid_capital.savings_account_balance or 0,
        plan.liquid_capital.idle_cash_for_investment or 0,
        plan.liquid_capital.fd_breakable_for_investment or 0,
        plan.liquid_capital.bonus_expected_for_investment or 0,
    ])

    from datetime import datetime
    age = int(pd_map.get("Age") or 30)
    plan.freedom_score_inputs = FreedomScoreInputs(
        age=age,
        monthly_income=monthly_income,
        monthly_expenses=me_total,
        monthly_emi=loan_emi + emi_other,
        portfolio_current_value=portfolio_total,
        liquid_assets_current_value=liquid_total,
        equity_allocation_percent=60,
    )

    return plan


def fmt_inr(n: float) -> str:
    """Indian-style ₹ formatter (lakh/crore)."""
    if n is None:
        return "—"
    n = float(n)
    if abs(n) >= 1e7:
        return f"₹{n/1e7:,.2f} Cr"
    if abs(n) >= 1e5:
        return f"₹{n/1e5:,.2f} L"
    return f"₹{n:,.0f}"


def run_persona(path: Path) -> None:
    print(f"\n{'='*70}")
    print(f"  PERSONA: {path.stem}")
    print(f"{'='*70}")

    plan = build_plan_from_excel(path)

    # ── basic profile ──────────────────────────────────────────────────
    print(f"\n  Profile: {plan.personal_details.full_name}, age "
          f"{plan.freedom_score_inputs.age}, retire at "
          f"{plan.personal_details.retirement_age_target}")
    print(f"  Income {fmt_inr(plan.freedom_score_inputs.monthly_income*12)}/yr | "
          f"Expenses {fmt_inr(plan.freedom_score_inputs.monthly_expenses*12)}/yr | "
          f"EMI {fmt_inr(plan.freedom_score_inputs.monthly_emi*12)}/yr")
    print(f"  Portfolio {fmt_inr(plan.freedom_score_inputs.portfolio_current_value)} | "
          f"Liquid {fmt_inr(plan.freedom_score_inputs.liquid_assets_current_value)} | "
          f"NFA {fmt_inr(sum(r.current_value or 0 for r in plan.real_estate) + sum(g.current_value or 0 for g in plan.gold))}")

    # ── CFP engine ────────────────────────────────────────────────────
    out = compute_cfp(plan)
    s = out.summary

    print(f"\n  ── CFP SUMMARY ────────────────────────────────────────────────")
    print(f"    Current age:            {s['current_age']}")
    print(f"    Retirement age:         {s['retirement_age']}")
    print(f"    Life expectancy:        {s['life_expectancy']}")
    print(f"    Gross savings rate:     {s['gross_savings_rate']*100:.1f}%")
    print(f"    Required savings rate:  {s['required_savings_rate']*100:.1f}%")
    print(f"    On track:               {s['on_track']}")
    print(f"    Total required SIP/mo:  {fmt_inr(s['total_required_sip_monthly'])}")
    print(f"    Total incremental SIP:  {fmt_inr(s['total_incremental_sip_monthly'])}")
    print(f"    Horizon NW estimate:    {fmt_inr(s['horizon_net_worth_estimate'])}")
    print(f"    Holdings-wt'd ROI:      {s['fa_holdings_weighted_roi']*100:.2f}%")

    # ── Goal-by-goal comparison vs Excel FV ───────────────────────────
    # Caveat: Excel's stored FV column was computed when current year=2025;
    # the engine uses datetime.now() = 2026. Compare engine vs Excel by
    # recomputing what the Excel formula `today_cost*(1+infl)^n` SHOULD
    # produce at the engine's current horizon, then verifying the engine
    # matches that. This isolates engine correctness from baseline drift.
    from datetime import datetime
    engine_current_yr = datetime.now().year
    print(f"\n  ── GOALS — Engine FV vs Excel formula (n={'{target_year} - '+str(engine_current_yr)}) ──")
    wb = load_workbook(path, data_only=True)
    ws_goals = wb["10_Financial_Goals"]
    excel_data = {}
    for r in range(2, ws_goals.max_row + 1):
        name = ws_goals.cell(r, 1).value
        ty = ws_goals.cell(r, 2).value
        tc = ws_goals.cell(r, 3).value
        infl_raw = ws_goals.cell(r, 4).value
        fv_excel = ws_goals.cell(r, 5).value
        if name and tc and ty and "total" not in str(name).lower():
            infl = float(str(infl_raw).strip("% ")) / 100 if infl_raw else 0.07
            excel_data[str(name)] = {
                "today_cost": float(tc),
                "target_year": int(ty),
                "inflation": infl,
                "fv_excel_2025_base": float(fv_excel) if fv_excel else None,
            }
    for gb in out.goal_blocks:
        gname = gb.get("goal_name", "?")
        eng_fv = gb.get("future_value_needed", 0)
        eng_sip = gb.get("required_sip_monthly", 0)
        eng_n = gb.get("years_to_go", 0)
        ed = excel_data.get(gname)
        if ed:
            n = ed["target_year"] - engine_current_yr
            expected_fv = ed["today_cost"] * ((1 + ed["inflation"]) ** n)
            delta_pct = ((eng_fv - expected_fv) / expected_fv * 100) if expected_fv else None
            marker = "✓" if delta_pct is not None and abs(delta_pct) < 0.5 else ("≈" if delta_pct is not None and abs(delta_pct) < 2 else "✗")
            print(f"    {marker} {gname:16s}  TodayCost: {fmt_inr(ed['today_cost']):>11s}  "
                  f"Infl: {ed['inflation']*100:.0f}%  n={n:>2d}y  "
                  f"Eng FV: {fmt_inr(eng_fv):>11s}  Expected: {fmt_inr(expected_fv):>11s}  "
                  f"Δ {f'{delta_pct:+.3f}%' if delta_pct is not None else '—':>9s}  "
                  f"SIP {fmt_inr(eng_sip)}/mo")
            if ed["fv_excel_2025_base"]:
                ratio = eng_fv / ed["fv_excel_2025_base"]
                # Sanity: engine should be ≈ excel × (1+infl)^-1 if 1-year baseline drift
                expected_drift = 1 / (1 + ed["inflation"])
                drift_ok = abs(ratio - expected_drift) / expected_drift < 0.02
                print(f"        Excel(2025-base) FV: {fmt_inr(ed['fv_excel_2025_base'])}  → engine/excel ratio={ratio:.4f}  "
                      f"(expected ~{expected_drift:.4f} for 1y baseline drift)  {'✓ matches drift' if drift_ok else '✗ off'}")

    # ── Retirement detail ─────────────────────────────────────────────
    ret = out.retirement
    print(f"\n  ── RETIREMENT ─────────────────────────────────────────────────")
    print(f"    Corpus required:        {fmt_inr(ret.get('corpus_required'))}")
    print(f"    Existing FV (EPF/PPF):  {fmt_inr(ret.get('existing_retirement_assets_fv', 0))}")
    print(f"    Shortfall:              {fmt_inr(ret.get('corpus_shortfall_after_existing', ret.get('corpus_required')))}")
    print(f"    Required SIP/mo:        {fmt_inr(ret.get('required_monthly_sip', 0))}")

    # ── Insurance ─────────────────────────────────────────────────────
    ins = out.insurance
    print(f"\n  ── INSURANCE ──────────────────────────────────────────────────")
    print(f"    HLV cover required:     {fmt_inr(ins.get('hlv_cover_required'))}")
    print(f"    Needs cover required:   {fmt_inr(ins.get('needs_cover_required'))}")
    print(f"    Avg life cover req'd:   {fmt_inr(ins.get('avg_cover_required', ins.get('life_cover_required')))}")
    print(f"    Existing life cover:    {fmt_inr(ins.get('existing_term_cover', 0))}")
    print(f"    Additional cover req:   {fmt_inr(ins.get('additional_cover_required'))}")
    health = ins.get("health", {})
    print(f"    Health cover required:  {fmt_inr(health.get('required'))}")
    print(f"    Existing health cover:  {fmt_inr(health.get('existing_cover', 0))}")
    print(f"    Additional health gap:  {fmt_inr(health.get('additional_cover_required', 0))}")
    rule = health.get("rule_used")
    if rule:
        print(f"    Rule: {rule}")

    # ── Debt block ────────────────────────────────────────────────────
    debt = out.debt
    ratios = debt["ratios"]
    strategies = debt["strategies"]
    print(f"\n  ── DEBT MGT (Finding 7) ──────────────────────────────────────")
    print(f"    DSCR: {ratios['dscr']} ({ratios['dscr_status']})")
    print(f"    DTI:  {ratios['dti']} ({ratios['dti_status']})")
    print(f"    DNI:  {ratios['dni']} ({ratios['dni_status']})")
    print(f"    Total debt outstanding: {fmt_inr(ratios['total_debt_outstanding'])}")
    if strategies["loans"]:
        print(f"    Avalanche order: {strategies['avalanche_order']}")
        print(f"    Snowball  order: {strategies['snowball_order']}")
        print(f"    Default strategy: {strategies['default_strategy']}")
    else:
        print(f"    (no active loans)")

    # ── Tax regime ────────────────────────────────────────────────────
    tax = out.tax_regime
    print(f"\n  ── TAX REGIME (Finding 6) — FY {tax['fy']} ────────────────────")
    print(f"    Annual gross income:   {fmt_inr(tax['annual_gross_income'])}")
    print(f"    Old regime total tax:  {fmt_inr(tax['old_regime']['total_tax'])} "
          f"(eff. {tax['old_regime']['effective_rate']*100:.2f}%)")
    print(f"      Deductions: 80C={fmt_inr(tax['old_regime']['deductions']['80C'])} "
          f"80CCD(1B)={fmt_inr(tax['old_regime']['deductions']['80CCD_1B'])} "
          f"80D={fmt_inr(tax['old_regime']['deductions']['80D'])} "
          f"24(b)={fmt_inr(tax['old_regime']['deductions']['24b'])}")
    print(f"    New regime total tax:  {fmt_inr(tax['new_regime']['total_tax'])} "
          f"(eff. {tax['new_regime']['effective_rate']*100:.2f}%)")
    print(f"    Recommended: {tax['recommended_regime'].upper()} "
          f"— saves {fmt_inr(tax['annual_savings_with_recommended'])}/yr")

    # ── Net worth from legacy engine for comparison ────────────────────
    cf = compute_cashflow(plan, 30)
    print(f"\n  ── LEGACY CASHFLOW (Finding 1 unification check) ────────────")
    print(f"    Year-0 NW:              {fmt_inr(cf.rows[0].total_net_worth if cf.rows else 0)}")
    print(f"    Year-30 NW:             {fmt_inr(cf.rows[-1].total_net_worth if cf.rows else 0)}")
    print(f"    CFP horizon NW:         {fmt_inr(s['horizon_net_worth_estimate'])}")

    # ── Finding 1 — verify recompute() populates plan.computed.cfp ─────
    from stackwealth.skills.scenario import recompute, _to_dict, _from_dict
    plan_d = _to_dict(plan)
    plan_d = recompute(plan_d)
    cfp_snap = plan_d.get("computed", {}).get("cfp")
    if cfp_snap:
        snap_keys = list(cfp_snap.keys())
        snap_summary = cfp_snap.get("summary", {})
        print(f"\n  ── UNIFIED SNAPSHOT — plan.computed.cfp populated ✓ ─────────")
        print(f"    Keys: {snap_keys}")
        print(f"    Same retirement corpus?  {snap_summary.get('retirement_corpus_required') == s.get('retirement_corpus_required')}")
        print(f"    Same DSCR?               {snap_summary.get('dscr') == s.get('dscr')}")
        print(f"    Same tax regime?         {snap_summary.get('recommended_tax_regime') == s.get('recommended_tax_regime')}")
        nw = plan_d.get("computed", {}).get("net_worth", {})
        print(f"    NW snapshot total:       {fmt_inr(nw.get('total'))} (assets {fmt_inr(nw.get('assets_total'))} − unsecured debts {fmt_inr(nw.get('debts_total'))})")
    else:
        print(f"\n  ✗ UNIFIED SNAPSHOT — plan.computed.cfp NOT populated (Finding 1 broken)")


def main():
    repo_root = Path(__file__).resolve().parent.parent
    personas = [
        repo_root / "V1_Young_Professional_Rohan.xlsx",
        repo_root / "V2_Married_Family_Anjali_Vikram.xlsx",
        repo_root / "V3_Established_Professional_Suresh.xlsx",
        repo_root / "V4_Pre_Retirement_Rajeev.xlsx",
    ]
    for p in personas:
        if not p.exists():
            print(f"MISSING: {p}")
            continue
        try:
            run_persona(p)
        except Exception as e:
            import traceback
            print(f"\n  ✗ FAILED for {p.stem}: {e}")
            traceback.print_exc()

    # ── Questionnaire smoke test (Finding 9) ───────────────────────────
    print(f"\n{'='*70}")
    print(f"  RISK QUESTIONNAIRE (Finding 9) — bucketing sanity check")
    print(f"{'='*70}")
    for label, val in [("all-1 (Conservative)", 1), ("all-2", 2), ("all-3 (Moderate)", 3),
                       ("all-4", 4), ("all-5 (Aggressive)", 5)]:
        ans = {q["id"]: val for q in RISK_QUESTIONNAIRE}
        res = compute_questionnaire_score(ans)
        print(f"  {label:30s} total={res['total_score']:3d}  norm={res['normalised_score_0_100']:>5.1f}  "
              f"profile={res['recommended_profile']}")


if __name__ == "__main__":
    main()
